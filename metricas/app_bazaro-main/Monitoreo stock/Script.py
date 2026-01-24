"""
Sistema de Monitoreo de Stock para Mercado Libre Chile
Analiza ventas de últimos 30 días y calcula reposición por tipo de logística
Agrupa publicaciones duplicadas por SKU y calcula cajas master
"""

import requests
import pandas as pd
from datetime import datetime, timedelta
import json
import webbrowser
from urllib.parse import urlencode
import hashlib
import base64
import secrets
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment

class MercadoLibreStockMonitor:
    """Monitor de stock basado en ventas de últimos 30 días"""
    
    def __init__(self, app_id, client_secret, excel_file_path):
        self.app_id = app_id
        self.client_secret = client_secret
        self.excel_file = excel_file_path
        self.access_token = None
        self.refresh_token = None
        self.user_id = None
        self.base_url = "https://api.mercadolibre.com"
        self.auth_url = "https://auth.mercadolibre.cl"
        self.redirect_uri = "https://www.google.com"
        self.code_verifier = None
        self.code_challenge = None
        
        # ==================== OPTIMIZACIONES ====================
        # Sistema de Caché en Memoria
        self._item_cache = {}
        self._order_cache = {}
        self._shipping_cache = {}
        self._stock_cache = {}
        self._inventory_cache = {}  # Nueva caché para inventory IDs
        self._transit_cache = {}     # Nueva caché para unidades en tránsito
        
        # Thread Safety
        import threading
        self._cache_lock = threading.Lock()
        
        # Estadísticas de Caché
        self._cache_hits = 0
        self._cache_misses = 0

    # ==================== AUTENTICACIÓN ====================
    
    def _generate_pkce_codes(self):
        """Genera códigos PKCE para seguridad adicional"""
        self.code_verifier = base64.urlsafe_b64encode(
            secrets.token_bytes(32)
        ).decode('utf-8').rstrip('=')
        
        challenge_bytes = hashlib.sha256(
            self.code_verifier.encode('utf-8')
        ).digest()
        self.code_challenge = base64.urlsafe_b64encode(
            challenge_bytes
        ).decode('utf-8').rstrip('=')
    
    def get_authorization_url(self):
        """Genera URL de autorización y abre navegador"""
        self._generate_pkce_codes()
        
        params = {
            'response_type': 'code',
            'client_id': self.app_id,
            'redirect_uri': self.redirect_uri,
            'code_challenge': self.code_challenge,
            'code_challenge_method': 'S256'
        }
        
        auth_url = f"{self.auth_url}/authorization?{urlencode(params)}"
        
        print("\n" + "="*80)
        print("PASO 1: AUTORIZACION")
        print("="*80)
        print("\nAbriendo navegador...")
        print("\nEn el navegador:")
        print("  1. Inicia sesion con tu cuenta de Mercado Libre Chile")
        print("  2. Autoriza la aplicacion")
        print("  3. Seras redirigido a Google.com (mostrara error - es normal)")
        print("  4. En la barra de direcciones veras:")
        print("     https://www.google.com/?code=TG-123456...")
        print("  5. Copia TODO el codigo despues de '?code='")
        print("="*80)
        
        try:
            webbrowser.open(auth_url)
        except:
            print("\nNo se pudo abrir el navegador. Copia esta URL:")
            print(f"\n{auth_url}\n")
        
        return auth_url
    
    def get_access_token(self, authorization_code):
        """Intercambia código por access token"""
        url = f"{self.base_url}/oauth/token"
        
        data = {
            'grant_type': 'authorization_code',
            'client_id': self.app_id,
            'client_secret': self.client_secret,
            'code': authorization_code,
            'redirect_uri': self.redirect_uri,
            'code_verifier': self.code_verifier
        }
        
        headers = {
            'accept': 'application/json',
            'content-type': 'application/x-www-form-urlencoded'
        }
        
        try:
            response = requests.post(url, data=data, headers=headers)
            response.raise_for_status()
            
            token_data = response.json()
            self.access_token = token_data['access_token']
            self.refresh_token = token_data['refresh_token']
            self.user_id = token_data['user_id']
            
            self._save_tokens(token_data)
            
            print("\nToken obtenido exitosamente")
            print(f"User ID: {self.user_id}")
            
            return token_data
            
        except requests.exceptions.RequestException as e:
            print(f"\nError al obtener token: {e}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    print(f"Detalles: {e.response.json()}")
                except:
                    print(f"Detalles: {e.response.text}")
            return None
    
    def _save_tokens(self, token_data):
        """Guarda tokens en archivo"""
        token_data['timestamp'] = datetime.now().isoformat()
        with open('ml_tokens.json', 'w') as f:
            json.dump(token_data, f, indent=2)
    
    def load_tokens(self):
        """Carga tokens guardados"""
        try:
            with open('ml_tokens.json', 'r') as f:
                token_data = json.load(f)
                self.access_token = token_data.get('access_token')
                self.refresh_token = token_data.get('refresh_token')
                self.user_id = token_data.get('user_id')
                return True
        except FileNotFoundError:
            return False
    
    def get_user_info(self):
        """Obtiene información del usuario autenticado"""
        url = f"{self.base_url}/users/me"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            user_data = response.json()
            self.user_id = user_data['id']
            print(f"Usuario: {user_data.get('nickname', 'N/A')}")
            print(f"User ID: {self.user_id}")
            return user_data
        except:
            return None
    
    # ==================== OBTENER PUBLICACIONES ====================
    
    def get_all_items(self, only_active=False):
        """Obtiene todas las publicaciones (activas e inactivas)
        
        Args:
            only_active: Si es True, solo retorna publicaciones activas
        """
        print("\nObteniendo publicaciones...")
        
        all_items = []
        
        # Activas
        active = self._fetch_items_by_status('active')
        all_items.extend(active)
        print(f"  Activas: {len(active)}")
        
        if only_active:
            print(f"\n⚡ FILTRO ACTIVO: Solo analizando productos activos")
            print(f"Total: {len(all_items)} publicaciones (pausadas y cerradas excluidas)")
            return all_items
        
        # Pausadas (sin stock)
        paused = self._fetch_items_by_status('paused')
        all_items.extend(paused)
        print(f"  Pausadas: {len(paused)}")
        
        # Cerradas
        closed = self._fetch_items_by_status('closed')
        all_items.extend(closed)
        print(f"  Cerradas: {len(closed)}")
        
        print(f"\nTotal: {len(all_items)} publicaciones")
        
        return all_items
    
    def _fetch_items_by_status(self, status):
        """Obtiene items por estado"""
        url = f"{self.base_url}/users/{self.user_id}/items/search"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        items = []
        offset = 0
        limit = 50
        
        try:
            while True:
                params = {
                    'status': status,
                    'offset': offset,
                    'limit': limit
                }
                
                response = requests.get(url, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()
                
                batch = data.get('results', [])
                if not batch:
                    break
                
                items.extend(batch)
                offset += limit
                
                if len(batch) < limit:
                    break
                    
        except Exception as e:
            print(f"Error obteniendo items {status}: {e}")
        
        return items
    
    def get_item_details(self, item_id):
        """Obtiene los detalles completos de un item (con caché)"""
        # Verificar caché primero
        with self._cache_lock:
            if item_id in self._item_cache:
                self._cache_hits += 1
                return self._item_cache[item_id]
        
        # No está en caché, hacer request
        self._cache_misses += 1
        url = f"{self.base_url}/items/{item_id}"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()
            
            # Guardar en caché
            with self._cache_lock:
                self._item_cache[item_id] = data
            
            return data
        except Exception as e:
            print(f"Error obteniendo detalles del item {item_id}: {e}")
            return None
    
    # ==================== NUEVAS FUNCIONES PARA UNIDADES EN TRÁNSITO ====================
    
    def get_inventory_id(self, item_id):
        """Obtiene el inventory_id de una publicación"""
        # Verificar caché primero
        with self._cache_lock:
            if item_id in self._inventory_cache:
                return self._inventory_cache[item_id]
        
        # Obtener detalles del item
        item_details = self.get_item_details(item_id)
        if not item_details:
            return None
        
        inventory_id = item_details.get('inventory_id')
        
        # Guardar en caché
        with self._cache_lock:
            self._inventory_cache[item_id] = inventory_id
        
        return inventory_id
    
    def get_transit_stock(self, item_id):
        """Obtiene las unidades en tránsito a FULL para un producto
        
        Returns:
            int: Cantidad de unidades en tránsito a FULL
        """
        # Verificar caché primero
        with self._cache_lock:
            if item_id in self._transit_cache:
                return self._transit_cache[item_id]
        
        # Obtener inventory_id
        inventory_id = self.get_inventory_id(item_id)
        if not inventory_id:
            with self._cache_lock:
                self._transit_cache[item_id] = 0
            return 0
        
        # Consultar stock en fulfillment
        url = f"{self.base_url}/inventories/{inventory_id}/stock/fulfillment"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            data = response.json()

            
            transit_quantity = 0
            
            # La respuesta puede ser una lista o un diccionario
            if isinstance(data, list):
                for stock_item in data:
                    # Verificar que sea un diccionario
                    if isinstance(stock_item, dict):
                        status = stock_item.get('status')
                        quantity = stock_item.get('quantity', 0)
                        
                        if status == 'transfer':
                            transit_quantity += quantity
            elif isinstance(data, dict):
                # Si es un diccionario, buscar en posibles estructuras
                items = data.get('items', []) or data.get('stock', [])
                for stock_item in items:
                    if isinstance(stock_item, dict):
                        status = stock_item.get('status')
                        quantity = stock_item.get('quantity', 0)
                        
                        if status == 'transfer':
                            transit_quantity += quantity
            
            # Guardar en caché
            with self._cache_lock:
                self._transit_cache[item_id] = transit_quantity
            
            return transit_quantity
            
        except requests.exceptions.RequestException as e:
            # Si hay error (ej: item sin FULL), retornar 0
            with self._cache_lock:
                self._transit_cache[item_id] = 0
            return 0
        except Exception as e:
            # Cualquier otro error
            with self._cache_lock:
                self._transit_cache[item_id] = 0
            return 0
    
    # ==================== ANÁLISIS DE VENTAS ====================
    
    def get_sales_last_30_days(self, item_id):
        """Obtiene ventas de últimos 30 días por tipo de logística"""
        sales = {'flex': 0, 'full': 0, 'centro_envio': 0}
        dict_info = {'costo_envio': 0, 'precio_unitario': 0, 'comision': 0, 'fecha_ultima_venta': None}
        
        date_from = (datetime.now() - timedelta(days=30)).strftime(
            '%Y-%m-%dT00:00:00.000-04:00'
        )
        
        url = f"{self.base_url}/orders/search"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        params = {
            'seller': self.user_id,
            'item': item_id,
            'order.date_created.from': date_from
        }
        
        try:
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            results = data.get('results', [])
            print(f"  {item_id}: {len(results)} órdenes encontradas")

            if len(results) != 0:
                # Obtener información de la última orden (más reciente)
                ultima_orden = results[-1]
                order_items = ultima_orden.get('order_items')
                precio_unitario = order_items[0]['unit_price']
                comision = order_items[0]['sale_fee']
                
                # Capturar fecha de la última venta
                fecha_ultima_venta = ultima_orden.get('date_created')

                shipping_id = ultima_orden.get('shipping', {}).get('id')
                shipping_info = self._get_costs_info(shipping_id)
                costo_envio = shipping_info.get('senders')[0].get('cost')
                dict_info = {'costo_envio': costo_envio,
                            'precio_unitario': precio_unitario,
                            'comision': comision,
                            'fecha_ultima_venta': fecha_ultima_venta}
                print(f'Información order: {dict_info}')
            
            for order_data in results:
                payments = order_data.get('payments')
                order_id = payments[0]['order_id']
                print(f"    Procesando orden: {order_id}")
                
                order_detail = self._get_order_detail(order_id)
                if not order_detail:
                    continue
                
                order_status = order_detail.get('status', '')
                if order_status != 'paid':
                    continue
                
                quantity = self._get_item_quantity(order_detail, item_id)
                if quantity == 0:
                    continue
                
                shipping_id = order_detail.get('shipping', {}).get('id')
                if shipping_id:
                    shipping_info = self._get_shipping_info(shipping_id)
                    if shipping_info:
                        logistic_type = shipping_info.get('logistic_type', '')
                        print(f"    Orden {order_id}: qty={quantity}, logistic_type={logistic_type}")
                        
                        if logistic_type == 'fulfillment':
                            sales['full'] += quantity
                        elif logistic_type == 'xd_drop_off':
                            sales['centro_envio'] += quantity
                        else:
                            sales['flex'] += quantity
                    else:
                        sales['flex'] += quantity
                else:
                    sales['flex'] += quantity
            
            if sales['flex'] > 0 or sales['full'] > 0 or sales['centro_envio'] > 0:
                print(f"    -> TOTAL FLEX: {sales['flex']}, FULL: {sales['full']}, CENTRO ENVIO: {sales['centro_envio']}")
                        
        except Exception as e:
            print(f"Error obteniendo ventas para {item_id}: {e}")
        
        return sales, dict_info

    def _get_shipping_info(self, shipping_id):
        """Obtiene información del envío"""
        url = f"{self.base_url}/shipments/{shipping_id}"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except:
            return None
        
    def _get_costs_info(self, shipping_id):
        """Obtiene informaciÃ³n del envÃ­o"""
        url = f"{self.base_url}/shipments/{shipping_id}/costs"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except:
            return None

    
    def _is_santiago_rm(self, shipping_info):
        """Determina si el envío es a Santiago o Región Metropolitana"""
        if not shipping_info:
            return False
        
        try:
            # Obtener la dirección del receptor (destino del envío)
            receiver_address = shipping_info.get('receiver_address', {})
            
            city_obj = receiver_address.get('city', {})
            state_obj = receiver_address.get('state', {})
            neighborhood_obj = receiver_address.get('neighborhood', {})
            
            city = city_obj.get('name', '').lower() if city_obj else ''
            state = state_obj.get('name', '').lower() if state_obj else ''
            state_id = state_obj.get('id', '').lower() if state_obj else ''
            neighborhood = neighborhood_obj.get('name', '').lower() if neighborhood_obj else ''
            
            # Verificar si es Santiago o Región Metropolitana
            santiago_keywords = ['santiago', 'stgo', 'providencia', 'las condes', 'vitacura', 
                               'la reina', 'ñuñoa', 'macul', 'peñalolén', 'la florida',
                               'san miguel', 'san joaquín', 'la cisterna', 'lo espejo',
                               'pedro aguirre cerda', 'san ramón', 'el bosque', 'la granja',
                               'san bernardo', 'puente alto', 'maipú', 'quilicura',
                               'huechuraba', 'recoleta', 'independencia', 'conchalí',
                               'renca', 'quinta normal', 'lo prado', 'pudahuel', 'cerro navia',
                               'estación central', 'cerrillos', 'colina', 'lampa', 'til til',
                               'pirque', 'san josé de maipo', 'buin', 'paine', 'melipilla',
                               'talagante', 'peñaflor', 'isla de maipo', 'el monte',
                               'padre hurtado', 'curacaví']
            
            rm_keywords = ['metropolitana', 'rm', 'región metropolitana', 'cl-rm']
            
            # Verificar ciudad
            for keyword in santiago_keywords:
                if keyword in city or keyword in neighborhood:
                    return True
            
            # Verificar región (incluye verificar state_id)
            for keyword in rm_keywords:
                if keyword in state or keyword in state_id:
                    return True
            
            return False
        except Exception as e:
            return False
    
    def _is_colina(self, shipping_info):
        """Determina si el envío es específicamente a Colina (dirección del cliente/receptor)"""
        if not shipping_info:
            return False
            
        try:
            # SOLO verificar dirección del receptor (cliente/destino)
            # NO verificar sender_address (vendedor/bodega)
            receiver_address = shipping_info.get('receiver_address', {})
            receiver_city = receiver_address.get('city', {}).get('name', '').lower()
            receiver_neighborhood = receiver_address.get('neighborhood', {}).get('name', '').lower()
            
            # Es Colina solo si el CLIENTE está en Colina
            colina_keywords = ['colina']
            
            for keyword in colina_keywords:
                if keyword in receiver_city or keyword in receiver_neighborhood:
                    return True
            
            return False
        except Exception as e:
            return False

    def analyze_flex_viability(self, test_mode=False, test_limit=50, only_active=False):
        """Analiza la viabilidad de Flex para cada producto"""
        print("\n" + "="*80)
        if test_mode:
            print(f"ANÁLISIS DE VIABILIDAD FLEX - MODO PRUEBA ({test_limit} items)")
        else:
            print("ANÁLISIS DE VIABILIDAD FLEX POR PRODUCTO")
        if only_active:
            print("⚡ FILTRO ACTIVADO - Solo productos activos")
        print("="*80)
        
        # Obtener todas las publicaciones
        all_items = self.get_all_items(only_active=only_active)
        
        # Aplicar límite en modo test
        if test_mode:
            all_items = all_items[:test_limit]
            print(f"\n✓ MODO PRUEBA: Analizando solo {len(all_items)} publicaciones")
        
        # Cargar Excel con información de SKU y proveedores
        df_excel = None
        item_id_column = None
        sku_column = None
        proveedor_column = None
        
        try:
            # Leer la hoja "Publicaciones"
            df_excel = pd.read_excel(self.excel_file, sheet_name='Publicaciones')
            
            # La primera fila contiene los nombres reales de las columnas
            if len(df_excel) > 0:
                df_excel.columns = df_excel.iloc[0]
                df_excel = df_excel[1:].reset_index(drop=True)
            
            print(f"\n✓ Excel cargado correctamente")
            print(f"  Registros: {len(df_excel)}")
            print(f"  Columnas: {list(df_excel.columns)}")
            
            # Buscar columna de Item ID
            possible_item_columns = ['Número de publicación', 'ITEM_ID', 'Item ID', 'item_id', 'ID', 'ItemID']
            for col in possible_item_columns:
                if col in df_excel.columns:
                    item_id_column = col
                    print(f"  ✓ Item ID: '{col}'")
                    break
            
            # Buscar columna de SKU
            possible_sku_columns = ['SKU', 'sku', 'Sku', 'SKU_ITEM']
            for col in possible_sku_columns:
                if col in df_excel.columns:
                    sku_column = col
                    print(f"  ✓ SKU: '{col}'")
                    break
            
            # Buscar columna de Proveedor
            possible_prov_columns = ['Proveedor', 'proveedor', 'PROVEEDOR', 'Supplier', 'Provider']
            for col in possible_prov_columns:
                if col in df_excel.columns:
                    proveedor_column = col
                    print(f"  ✓ Proveedor: '{col}'")
                    break
            
            if not item_id_column:
                print("  ⚠ ADVERTENCIA: No se encontró columna de Item ID")
                print("    El análisis continuará sin datos de SKU y Proveedor")
                df_excel = None
                
        except Exception as e:
            print(f"⚠ Error al cargar Excel: {e}")
            print("  El análisis continuará sin datos de SKU y Proveedor")
            df_excel = None
        
        flex_analysis_data = []
        
        for idx, item_id in enumerate(all_items, 1):
            # item_id es directamente el string del ID
            # Obtener detalles del item
            item_details = self.get_item_details(item_id)
            if not item_details:
                print(f"Error obteniendo detalles de {item_id}")
                continue
            
            title = item_details.get('title', 'N/A')
            
            print(f"\n[{idx}/{len(all_items)}] Analizando: {item_id} - {title[:50]}...")
            
            # Buscar SKU y Proveedor en Excel
            sku = 'N/A'
            proveedor = 'N/A'
            
            if df_excel is not None and item_id_column is not None:
                try:
                    # Convertir item_id a string para la comparación
                    item_row = df_excel[df_excel[item_id_column].astype(str) == str(item_id)]
                    
                    if not item_row.empty:
                        if sku_column and sku_column in df_excel.columns:
                            sku_val = item_row.iloc[0][sku_column]
                            sku = str(sku_val) if pd.notna(sku_val) else 'N/A'
                        if proveedor_column and proveedor_column in df_excel.columns:
                            prov_val = item_row.iloc[0][proveedor_column]
                            proveedor = str(prov_val) if pd.notna(prov_val) else 'N/A'
                except Exception as e:
                    pass  # Continuar aunque falle la búsqueda en Excel
            
            # Obtener órdenes de los últimos 30 días
            date_from = (datetime.now() - timedelta(days=30)).strftime(
                '%Y-%m-%dT00:00:00.000-04:00'
            )
            
            url = f"{self.base_url}/orders/search"
            headers = {'Authorization': f'Bearer {self.access_token}'}
            params = {
                'seller': self.user_id,
                'item': item_id,
                'order.date_created.from': date_from
            }
            
            # Contadores por tipo de logística y ubicación
            stats = {
                'flex_santiago': 0,
                'flex_fuera': 0,
                'full_santiago': 0,
                'full_fuera': 0,
                'centro_santiago': 0,
                'centro_fuera': 0,
                'colina_total': 0  # Nuevo contador para Colina
            }
            
            try:
                response = requests.get(url, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()
                
                results = data.get('results', [])
                print(f"  Órdenes encontradas: {len(results)}")
                
                for order_data in results:
                    try:
                        payments = order_data.get('payments')
                        if not payments:
                            continue
                        
                        order_id = payments[0]['order_id']
                        
                        # Obtener detalle de la orden
                        order_detail = self._get_order_detail(order_id)
                        if not order_detail:
                            continue
                        
                        # Verificar que esté pagada
                        order_status = order_detail.get('status', '')
                        if order_status != 'paid':
                            continue
                        
                        # Obtener cantidad del producto
                        quantity = self._get_item_quantity(order_detail, item_id)
                        if quantity == 0:
                            continue
                        
                        # Obtener información del envío
                        shipping_id = order_detail.get('shipping', {}).get('id')
                        if not shipping_id:
                            continue
                        
                        shipping_info = self._get_shipping_info(shipping_id)
                        if not shipping_info:
                            continue
                        
                        # Determinar tipo de logística desde logistic.type
                        logistic = shipping_info.get('logistic', {})
                        logistic_type = logistic.get('type', 'flex')
                        
                        # Determinar ubicación (Santiago/RM o Fuera)
                        is_santiago = self._is_santiago_rm(shipping_info)
                        location = 'santiago' if is_santiago else 'fuera'
                        
                        # Verificar si es Colina (para Santiago solamente)
                        is_colina = self._is_colina(shipping_info)
                        if is_santiago and is_colina:
                            stats['colina_total'] += quantity
                        
                        # Clasificar según tipo de logística
                        # fulfillment = FULL
                        # drop_off = Centro de Envío
                        # otros = Flex
                        if logistic_type == 'fulfillment':
                            stats[f'full_{location}'] += quantity
                        elif logistic_type == 'drop_off':
                            stats[f'centro_{location}'] += quantity
                        else:
                            stats[f'flex_{location}'] += quantity
                        
                    except Exception as e:
                        continue
                
                # Calcular totales
                total_flex = stats['flex_santiago'] + stats['flex_fuera']
                total_full = stats['full_santiago'] + stats['full_fuera']
                total_centro = stats['centro_santiago'] + stats['centro_fuera']
                total_santiago = stats['flex_santiago'] + stats['full_santiago'] + stats['centro_santiago']
                total_fuera = stats['flex_fuera'] + stats['full_fuera'] + stats['centro_fuera']
                total_general = total_flex + total_full + total_centro
                
                # Calcular porcentajes
                pct_santiago = (total_santiago / total_general * 100) if total_general > 0 else 0
                pct_fuera = (total_fuera / total_general * 100) if total_general > 0 else 0
                pct_colina = (stats['colina_total'] / total_santiago * 100) if total_santiago > 0 else 0
                
                # Determinar recomendación
                recomendacion = 'SIN DATOS'
                if total_general > 0:
                    if pct_santiago >= 70:
                        recomendacion = 'ACTIVAR FLEX'
                    elif pct_santiago >= 50:
                        recomendacion = 'CONSIDERAR FLEX'
                    else:
                        recomendacion = 'NO RECOMENDADO'
                
                # Agregar a datos de análisis
                flex_analysis_data.append({
                    'Item ID': item_id,
                    'SKU': sku,
                    'Título': title,
                    'Proveedor': proveedor,
                    'Flex Santiago': stats['flex_santiago'],
                    'Flex Fuera': stats['flex_fuera'],
                    'Full Santiago': stats['full_santiago'],
                    'Full Fuera': stats['full_fuera'],
                    'Centro Santiago': stats['centro_santiago'],
                    'Centro Fuera': stats['centro_fuera'],
                    'Total Flex': total_flex,
                    'Total Full': total_full,
                    'Total Centro': total_centro,
                    'Total Santiago': total_santiago,
                    'Colina (de Santiago)': stats['colina_total'],  # Nueva columna
                    '% Colina (de Santiago)': pct_colina,  # Nueva columna
                    'Total Fuera': total_fuera,
                    'Total General': total_general,
                    '% Santiago': pct_santiago,
                    '% Fuera': pct_fuera,
                    'Recomendación': recomendacion
                })
                
                if total_general > 0:
                    colina_info = f" | Colina: {stats['colina_total']} ({pct_colina:.1f}% de Stgo)" if stats['colina_total'] > 0 else ""
                    print(f"  Total: {total_general} | Santiago: {total_santiago} ({pct_santiago:.1f}%){colina_info} | Fuera: {total_fuera} ({pct_fuera:.1f}%)")
                    print(f"  Recomendación: {recomendacion}")
                    
            except Exception as e:
                print(f"  Error analizando producto: {e}")
                continue
        
        # Crear DataFrame
        df_flex = pd.DataFrame(flex_analysis_data)
        
        # Filtrar solo productos con ventas
        df_flex = df_flex[df_flex['Total General'] > 0].copy()
        
        return df_flex

    def generate_flex_report(self, df_flex):
        """Genera reporte Excel de análisis Flex"""
        if df_flex is None or df_flex.empty:
            print("\nNo hay datos para generar reporte de análisis Flex")
            return
        
        from openpyxl.styles import PatternFill, Font, Alignment
        
        print("\n" + "="*80)
        print("GENERANDO REPORTE DE ANÁLISIS FLEX")
        print("="*80)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analisis_flex_{timestamp}.xlsx"
        
        # Crear resumen por recomendación
        summary_recomendacion = df_flex.groupby('Recomendación').agg({
            'Item ID': 'count',
            'Total General': 'sum',
            'Total Santiago': 'sum',
            'Total Fuera': 'sum',
            '% Santiago': 'mean'
        }).reset_index()
        
        summary_recomendacion.columns = ['Recomendación', 'Cantidad Productos', 
                                         'Total Ventas', 'Ventas Santiago', 
                                         'Ventas Fuera', '% Promedio Santiago']
        
        # Crear resumen por proveedor
        summary_proveedor = df_flex.groupby('Proveedor').agg({
            'Item ID': 'count',
            'Total General': 'sum',
            'Total Santiago': 'sum',
            'Total Fuera': 'sum',
            '% Santiago': 'mean'
        }).reset_index()
        
        summary_proveedor.columns = ['Proveedor', 'Cantidad Productos', 
                                      'Total Ventas', 'Ventas Santiago', 
                                      'Ventas Fuera', '% Promedio Santiago']
        
        # Ordenar por % Santiago descendente
        df_flex_sorted = df_flex.sort_values('% Santiago', ascending=False)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Hoja 1: Detalle completo
            df_flex_sorted.to_excel(writer, sheet_name='Detalle Análisis Flex', index=False)
            
            # Hoja 2: Resumen por recomendación
            summary_recomendacion.to_excel(writer, sheet_name='Resumen Recomendación', index=False)
            
            # Hoja 3: Resumen por proveedor
            summary_proveedor.to_excel(writer, sheet_name='Resumen Proveedor', index=False)
            
            # Hoja 4: Solo productos recomendados para Flex
            df_flex_recomendado = df_flex_sorted[df_flex_sorted['Recomendación'].isin(['ACTIVAR FLEX', 'CONSIDERAR FLEX'])]
            df_flex_recomendado.to_excel(writer, sheet_name='Recomendados Flex', index=False)
            
            # Formatear hojas
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Encabezados
                fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-ajustar columnas
                for col_idx, column in enumerate(worksheet.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Colores por recomendación en hoja Detalle
                if sheet_name == 'Detalle Análisis Flex':
                    for row_idx in range(2, len(df_flex_sorted) + 2):
                        recomendacion = worksheet.cell(row=row_idx, column=19).value  # Columna Recomendación
                        
                        if recomendacion == 'ACTIVAR FLEX':
                            color = '70AD47'  # Verde
                        elif recomendacion == 'CONSIDERAR FLEX':
                            color = 'FFD966'  # Amarillo
                        elif recomendacion == 'NO RECOMENDADO':
                            color = 'FF6B6B'  # Rojo
                        else:
                            color = 'D3D3D3'  # Gris
                        
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        for col in range(1, 20):  # 19 columnas
                            cell = worksheet.cell(row=row_idx, column=col)
                            cell.fill = fill
                    
                    # Formato porcentaje
                    for row in range(2, len(df_flex_sorted) + 2):
                        cell_q = worksheet.cell(row=row, column=17)  # % Santiago
                        cell_q.number_format = '0.00"%"'
                        
                        cell_r = worksheet.cell(row=row, column=18)  # % Fuera
                        cell_r.number_format = '0.00"%"'
        
        print(f"\nReporte de análisis Flex guardado: {filename}")
        print(f"Productos analizados: {len(df_flex)}")
        print(f"Productos recomendados para Flex: {len(df_flex_recomendado)}")

    def _get_order_detail(self, order_id):
        """Obtiene detalles de una orden"""
        url = f"{self.base_url}/orders/{order_id}"
        headers = {'Authorization': f'Bearer {self.access_token}'}
        
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except:
            return None
    
    def _get_item_quantity(self, order_detail, item_id):
        """Extrae cantidad del item en la orden"""
        for item in order_detail.get('order_items', []):
            if item.get('item', {}).get('id') == item_id:
                return item.get('quantity', 0)
        return 0
    
    # ==================== STOCK POR UBICACIÓN ====================
    
    def get_stock_by_location(self, item_id):
        """Obtiene stock separado por ubicación
        MODIFICADO: Incluye stock en tránsito a FULL
        Usa método correcto según documentación de MercadoLibre
        """
        stock = {
            'bodega': 0,
            'full': 0,
            'transito_full': 0,  # NUEVO: Unidades en tránsito a FULL
            'full_total': 0,     # NUEVO: FULL + Tránsito
            'is_supermarket': False
        }
        
        try:
            # Obtener datos del item usando caché
            item_data = self.get_item_details(item_id)
            if not item_data:
                return stock
            
            stock['is_supermarket'] = self.is_supermarket_item(item_data)
            
            # Obtener unidades en tránsito a FULL
            stock['transito_full'] = self.get_transit_stock(item_id)
            
            # Método 1: Usar inventories (más completo y confiable)
            inventories = item_data.get('inventories')
            if inventories:
                variations = inventories.get('variations', [])
                if variations:
                    for variation in variations:
                        available_qty = variation.get('available_quantity', 0)
                        location_ids = variation.get('location_ids', [])
                        
                        for location in location_ids:
                            loc_type = location.get('location_type', '')
                            
                            if loc_type == 'fulfillment':
                                stock['full'] += available_qty
                            elif loc_type in ['selling_address', 'seller_warehouse']:
                                if not stock['is_supermarket']:
                                    stock['bodega'] += available_qty
                    
                    # Calcular FULL total (stock + tránsito)
                    stock['full_total'] = stock['full'] + stock['transito_full']
                    return stock
            
            # Método 2: Fallback usando user_product_id
            user_product_id = item_data.get('user_product_id')
            if user_product_id:
                url = f"{self.base_url}/user-products/{user_product_id}/stock"
                headers = {'Authorization': f'Bearer {self.access_token}'}
                
                response = requests.get(url, headers=headers)
                if response.status_code == 200:
                    stock_data = response.json()
                    
                    for location in stock_data.get('locations', []):
                        loc_type = location.get('type', '')
                        quantity = location.get('quantity', 0)
                        
                        if loc_type == 'meli_facility':
                            stock['full'] += quantity
                        elif loc_type in ['selling_address', 'seller_warehouse']:
                            if not stock['is_supermarket']:
                                stock['bodega'] += quantity
                    
                    # Calcular FULL total (stock + tránsito)
                    stock['full_total'] = stock['full'] + stock['transito_full']
                    return stock
            
            # Método 3: Fallback simple usando available_quantity
            if stock['is_supermarket']:
                stock['bodega'] = 0
            else:
                stock['bodega'] = item_data.get('available_quantity', 0)
            
            # Calcular FULL total (stock + tránsito)
            stock['full_total'] = stock['full'] + stock['transito_full']
                
        except Exception as e:
            print(f"    ⚠️ Error obteniendo stock de {item_id}: {e}")
        
        return stock
        
    def is_supermarket_item(self, item_data):
        """Determina si un item es de supermercado"""
        tags = item_data.get('tags', [])
        if 'supermarket' in tags or 'supermarket_eligible' in tags:
            return True
        
        domain_id = item_data.get('domain_id', '')
        if 'SUPERMARKET' in domain_id.upper():
            return True
        
        category_id = item_data.get('category_id', '')
        supermarket_categories = ['MLC1403', 'MLC1430']
        if any(category_id.startswith(cat) for cat in supermarket_categories):
            return True
        
        return False
    
    # ==================== ANÁLISIS PRINCIPAL ====================
    
    def _analyze_single_product_stock(self, item_id, excel_dict, test_mode=False):
        """Analiza un solo producto para stock (thread-safe)
        MODIFICADO: Considera unidades en tránsito a FULL y captura fecha de última venta
        """
        try:
            stock = self.get_stock_by_location(item_id)
            sales, dict_info_order = self.get_sales_last_30_days(item_id)
            
            reponer_bodega = max(0, sales['flex'] + sales['centro_envio'] - stock['bodega'])
            # MODIFICADO: Usar full_total que incluye tránsito
            reponer_full = max(0, sales['full'] - stock['full_total'])
            
            if test_mode or reponer_bodega > 0 or reponer_full > 0:
                excel_info = excel_dict.get(item_id, {})
                
                precio_unitario = dict_info_order.get('precio_unitario', 0) or 0
                comision = dict_info_order.get('comision', 0) or 0
                costo_envio = dict_info_order.get('costo_envio', 0) or 0
                costo = excel_info.get('costo', 0) or 0
                utilidad = precio_unitario - comision - costo_envio - costo
                ventas_totales = int(sales['centro_envio']) + int(sales['flex']) + int(sales['full'])
                
                # Calcular última utilidad
                ultima_utilidad = 0
                if ventas_totales > 0 and precio_unitario > 0:
                    ultima_utilidad = precio_unitario - comision - costo_envio - costo
                else:
                    utilidad = 0
                
                # Calcular % de Rentabilidad
                porcentaje_rentabilidad = 0
                if costo > 0:
                    porcentaje_rentabilidad = (ultima_utilidad / costo) * 100
                
                unidades_caja = excel_info.get('unidades_caja_master', 1)
                total_reponer = int(reponer_bodega + reponer_full)
                cajas_master = round(total_reponer / unidades_caja, 1)
                
                # Extraer fecha de última venta
                fecha_ultima_venta = dict_info_order.get('fecha_ultima_venta')
                
                return {
                    'Item ID': item_id,
                    'SKU': excel_info.get('sku', ''),
                    'Titulo': excel_info.get('titulo', 'N/A')[:60],
                    'Proveedor': excel_info.get('proveedor', 'Sin datos'),
                    'Ventas CENTRO ENVIO 30d': int(sales['centro_envio']),
                    'Stock Bodega': int(stock['bodega']),
                    'Ventas FLEX 30d': int(sales['flex']),
                    'Reponer Bodega': int(reponer_bodega),
                    'Stock FULL': int(stock['full']),
                    'En Tránsito FULL': int(stock['transito_full']),  # NUEVO
                    'Stock FULL Total': int(stock['full_total']),      # NUEVO
                    'Ventas FULL 30d': int(sales['full']),
                    'Reponer FULL': int(reponer_full),
                    'Total Ventas 30d': int(sales['flex'] + sales['full'] + sales['centro_envio']),
                    'Total Reponer': total_reponer,
                    'Un x Caja Master': unidades_caja,
                    'Cajas Master': cajas_master,
                    'Costo Unitario': excel_info.get('costo', 0),
                    'Costo Total': total_reponer * excel_info.get('costo', 0),
                    'Precio Venta': precio_unitario,
                    'Comision': comision,
                    'Costo envio': costo_envio,
                    'Utilidad': utilidad,
                    'Ultima Utilidad': ultima_utilidad,
                    '% Rentabilidad': porcentaje_rentabilidad,
                    'Fecha Ultima Venta': fecha_ultima_venta  # NUEVO
                }
            return None
        except Exception as e:
            print(f"\n⚠️  Error en producto {item_id}: {e}")
            return None
    
    def analyze_stock_needs(self, test_mode=False, test_limit=50, use_parallel=True, only_active=False):
        """Analiza necesidades de reposición
        
        Args:
            test_mode: Si es True, analiza solo test_limit items
            test_limit: Número de items a analizar en modo test
            use_parallel: Si es True, usa procesamiento paralelo
            only_active: Si es True, solo analiza productos activos
        """
        print("\n" + "="*80)
        if test_mode:
            print(f"ANALISIS DE STOCK - MODO PRUEBA ({test_limit} items)")
        else:
            print("ANALISIS DE STOCK - ULTIMOS 30 DIAS")
        
        print("✨ NUEVO: Incluyendo análisis de unidades en tránsito a FULL")
        if use_parallel:
            print("⚡ PARALELIZACIÓN ACTIVADA - productos simultáneos")
        if only_active:
            print("⚡ FILTRO ACTIVADO - Solo productos activos")
        print("="*80)
        
        items = self.get_all_items(only_active=only_active)
        if not items:
            return None
        
        if test_mode:
            items = items[:test_limit]
            print(f"\nMODO PRUEBA: Analizando solo {len(items)} publicaciones")
        
        excel_data = self._load_excel()
        if excel_data is None:
            return None
        
        excel_dict = self._create_excel_dict(excel_data)
        
        results = []
        
        if use_parallel:
            # Procesamiento paralelo
            print(f"\n⚡ Iniciando análisis paralelo con workers...")
            with ThreadPoolExecutor(max_workers=12) as executor:
                # Crear futures para todos los productos
                future_to_item = {
                    executor.submit(self._analyze_single_product_stock, item_id, excel_dict, test_mode): item_id
                    for item_id in items
                }
                
                # Procesar resultados conforme se completan
                completed = 0
                for future in as_completed(future_to_item):
                    completed += 1
                    item_id = future_to_item[future]
                    
                    try:
                        result = future.result()
                        if result:
                            results.append(result)
                        
                        # Mostrar progreso cada 10 productos
                        if completed % 10 == 0 or completed == len(items):
                            print(f"  Progreso: {completed}/{len(items)} productos analizados")
                    
                    except Exception as e:
                        print(f"\n⚠️  Error procesando {item_id}: {e}")
        
        else:
            # Procesamiento secuencial (original)
            print("\nAnalizando ventas y stock...")
            for i, item_id in enumerate(items, 1):
                result = self._analyze_single_product_stock(item_id, excel_dict, test_mode)
                if result:
                    results.append(result)
                
                if i % 10 == 0 or i == len(items):
                    print(f"  Progreso: {i}/{len(items)} productos")
        
        # Mostrar estadísticas de caché
        total_cache_requests = self._cache_hits + self._cache_misses
        if total_cache_requests > 0:
            hit_rate = (self._cache_hits / total_cache_requests) * 100
            print(f"\n📊 Estadísticas de Caché:")
            print(f"  Hits: {self._cache_hits}")
            print(f"  Misses: {self._cache_misses}")
            print(f"  Hit Rate: {hit_rate:.1f}%")
            print(f"  Requests ahorrados: {self._cache_hits}")
        
        print("\n")
        
        if not results:
            print("No hay productos que necesiten reposicion")
            return pd.DataFrame()
        
        df = pd.DataFrame(results)
        
        # AGRUPAR POR SKU
        df_agrupado = self._agrupar_por_sku(df)
        df_agrupado = df_agrupado.sort_values('Proveedor')
        
        return df_agrupado
    
    def _agrupar_por_sku(self, df):
        """Agrupa publicaciones por SKU y consolida Item IDs
        MODIFICADO: Agrega columnas de tránsito y fecha de última venta
        """
        
        grouped = df.groupby(['SKU', 'Proveedor']).agg({
            'Item ID': lambda x: list(x),
            'Titulo': 'first',
            'Ventas CENTRO ENVIO 30d': 'sum',
            'Stock Bodega': 'max',
            'Ventas FLEX 30d': 'sum',
            'Reponer Bodega': 'sum',
            'Stock FULL': 'max',
            'En Tránsito FULL': 'max',      # NUEVO
            'Stock FULL Total': 'max',       # NUEVO
            'Ventas FULL 30d': 'sum',
            'Reponer FULL': 'sum',
            'Total Ventas 30d': 'sum',
            'Total Reponer': 'sum',
            'Un x Caja Master': 'first',
            'Cajas Master': 'sum',
            'Costo Unitario': 'first',
            'Costo Total': 'sum',
            'Precio Venta': 'first',
            'Comision': 'first',
            'Costo envio': 'first',
            'Utilidad': 'first',
            'Ultima Utilidad': 'first',
            '% Rentabilidad': 'first',
            'Fecha Ultima Venta': 'first'  # NUEVO
        }).reset_index()
        
        grouped['Item ID 1'] = grouped['Item ID'].apply(lambda x: x[0] if len(x) > 0 else '')
        grouped['Item ID 2'] = grouped['Item ID'].apply(lambda x: x[1] if len(x) > 1 else '')
        
        columnas_ordenadas = [
            'Item ID 1', 'Item ID 2', 'SKU', 'Titulo', 'Proveedor',
            'Ventas CENTRO ENVIO 30d', 'Stock Bodega', 'Ventas FLEX 30d', 'Reponer Bodega',
            'Stock FULL', 'En Tránsito FULL', 'Stock FULL Total',  # MODIFICADO: Agregadas columnas
            'Ventas FULL 30d', 'Reponer FULL',
            'Total Ventas 30d', 'Total Reponer', 'Un x Caja Master', 'Cajas Master',
            'Costo Unitario', 'Costo Total', 'Precio Venta', 'Comision', 'Costo envio', 'Utilidad', 
            'Ultima Utilidad', '% Rentabilidad', 'Fecha Ultima Venta'  # NUEVO
        ]
        
        grouped = grouped[columnas_ordenadas]
        grouped['Cajas Master'] = (grouped['Total Reponer'] / grouped['Un x Caja Master']).round(1)
        
        return grouped
    
    def _load_excel(self):
        """Carga datos del Excel"""
        print("\nCargando Excel...")
        try:
            df = pd.read_excel(
                self.excel_file,
                sheet_name='Publicaciones',
                skiprows=1
            )
            print(f"Excel cargado: {len(df)} productos")
            return df
        except Exception as e:
            print(f"Error cargando Excel: {e}")
            return None
    
    def _create_excel_dict(self, df):
        """Crea diccionario desde Excel"""
        excel_dict = {}
        
        for _, row in df.iterrows():
            item_id = str(row.get('Número de publicación', '')).strip()
            if item_id and item_id != 'nan':
                unidades_caja = row.get('un x Caja Maestra', 1)
                if pd.isna(unidades_caja) or unidades_caja == 0:
                    unidades_caja = 1
                
                excel_dict[item_id] = {
                    'proveedor': row.get('Proveedor', 'Sin proveedor'),
                    'sku': str(row.get('SKU', '')),
                    'titulo': str(row.get('Título', '')),
                    'costo': float(row.get('Costo', 0) or 0),
                    'unidades_caja_master': float(unidades_caja)
                }
        
        return excel_dict
    
    # ==================== REPORTE ====================
    
    def generate_report(self, df):
        """Genera reporte de reposición MEJORADO con ventas FULL/no FULL y hoja de Reposición FULL"""
        if df.empty:
            return
        
        from openpyxl.styles import PatternFill, Font, Alignment
        
        print("\n" + "="*80)
        print("REPORTE DE REPOSICION POR PROVEEDOR - MEJORADO")
        print("="*80)
        
        # Por proveedor
        for proveedor, group in df.groupby('Proveedor'):
            print(f"\n{'-'*80}")
            print(f"PROVEEDOR: {proveedor}")
            print(f"{'-'*80}")
            print(f"Items: {len(group)}")
            print(f"Unidades Bodega: {int(group['Reponer Bodega'].sum())}")
            print(f"Unidades FULL: {int(group['Reponer FULL'].sum())}")
            print(f"TOTAL ORDEN: {int(group['Total Reponer'].sum())} unidades")
            print(f"Costo: ${group['Costo Total'].sum():,.0f}")
            print(f"\nDetalle:")
            
            for _, row in group.iterrows():
                print(f"\n  {row['Titulo']}")
                item_ids = f"ID: {row['Item ID 1']}"
                if row['Item ID 2']:
                    item_ids += f" / {row['Item ID 2']}"
                print(f"  {item_ids} | SKU: {row['SKU']}")
                print(f"  BODEGA (FLEX/Centro Envio):")
                print(f"    - Stock actual: {row['Stock Bodega']}")
                print(f"    - Ventas 30d FLEX: {row['Ventas FLEX 30d']}")
                print(f"    - Ventas 30d CENTRO ENVIO: {row['Ventas CENTRO ENVIO 30d']}")
                print(f"    - A reponer: {row['Reponer Bodega']}")
                print(f"  FULL (Fulfillment ML):")
                print(f"    - Stock actual: {row['Stock FULL']}")
                print(f"    - Ventas 30d: {row['Ventas FULL 30d']}")
                print(f"    - A reponer: {row['Reponer FULL']}")
                print(f"  TOTAL VENTAS 30d: {row['Total Ventas 30d']} unidades")
                print(f"  CAJAS MASTER: {row['Cajas Master']} cajas ({row['Un x Caja Master']} un/caja)")
                print(f"  ORDEN COMPRA: {row['Total Reponer']} unid × ${row['Costo Unitario']:,.0f} = ${row['Costo Total']:,.0f}")
        
        print(f"\n{'='*80}")
        print("RESUMEN")
        print(f"{'='*80}")
        
        summary = df.groupby('Proveedor').agg({
            'SKU': 'count',
            'Ventas CENTRO ENVIO 30d': 'sum',
            'Ventas FLEX 30d': 'sum',
            'Ventas FULL 30d': 'sum',
            'Total Ventas 30d': 'sum',
            'Reponer Bodega': 'sum',
            'Reponer FULL': 'sum',
            'Total Reponer': 'sum',
            'Cajas Master': 'sum',
            'Costo Total': 'sum'
        }).reset_index()
        
        summary.columns = ['Proveedor', 'Items', 'Ventas CENTRO ENVIO', 'Ventas FLEX', 
                        'Ventas FULL', 'Tot Ventas', 'Rep Bodega', 'Rep FULL', 
                        'Tot Reponer', 'Tot Cajas Master', 'Costo']
        
        summary['Tot Cajas Master'] = summary['Tot Cajas Master'].round(1)
        
        print(summary.to_string(index=False))
        print(f"\nTotal ventas 30d: {int(df['Total Ventas 30d'].sum())} unidades")
        print(f"\nTotal a reponer: {int(df['Total Reponer'].sum())} unidades")
        print(f"Total cajas master: {df['Cajas Master'].sum():.1f} cajas")
        print(f"Inversion total: ${df['Costo Total'].sum():,.0f}")
        
        colores_proveedores = [
            'FFB6C1', 'B0E0E6', '98FB98', 'FFE4B5', 'DDA0DD',
            'F0E68C', 'FFA07A', '87CEEB', 'F5DEB3', 'D8BFD8',
        ]
        
        proveedores_unicos = df['Proveedor'].unique()
        color_map = {proveedor: colores_proveedores[i % len(colores_proveedores)] 
                    for i, proveedor in enumerate(proveedores_unicos)}
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_reposicion_{timestamp}.xlsx"
        
        proveedores_con_reposicion = []
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Detalle', index=False)
            summary.to_excel(writer, sheet_name='Resumen', index=False)
            
            # ==================== NUEVA HOJA: REPOSICIÓN FULL ====================
            df_full = df[df['Reponer FULL'] > 0].copy()
            
            if not df_full.empty:
                reposicion_full = df_full[['Item ID 1', 'Proveedor', 'SKU',
                                           'Stock FULL Total',  # ⭐ NUEVO: Stock actual en FULL
                                           'Ventas FULL 30d', 'Reponer FULL',
                                           'Costo Unitario', 'Ultima Utilidad']].copy()
                
                reposicion_full['Costo Reposición FULL'] = (reposicion_full['Reponer FULL'] * 
                                                             reposicion_full['Costo Unitario'])
                
                reposicion_full.columns = ['Código Publicación', 'Proveedor', 'SKU',
                                           'Stock FULL Actual',  # ⭐ NUEVO
                                           'Ventas FULL 30d', 'Unidades a Reponer',
                                           'Costo Unitario', 'Última Utilidad', 
                                           'Costo Reposición']
                
                reposicion_full = reposicion_full.sort_values(['Proveedor', 'Costo Reposición'], 
                                                              ascending=[True, False])
                
                total_full = pd.DataFrame({
                    'Código Publicación': ['TOTAL'],
                    'Proveedor': [''],
                    'SKU': [''],
                    'Stock FULL Actual': [int(reposicion_full['Stock FULL Actual'].sum())],  # ⭐ NUEVO
                    'Ventas FULL 30d': [int(reposicion_full['Ventas FULL 30d'].sum())],
                    'Unidades a Reponer': [int(reposicion_full['Unidades a Reponer'].sum())],
                    'Costo Unitario': [''],
                    'Última Utilidad': [reposicion_full['Última Utilidad'].mean()],
                    'Costo Reposición': [reposicion_full['Costo Reposición'].sum()]
                })
                
                reposicion_full = pd.concat([reposicion_full, total_full], ignore_index=True)
                reposicion_full.to_excel(writer, sheet_name='Reposición FULL', index=False)
                
                print(f"\n✅ Creada hoja 'Reposición FULL' con {len(df_full)} productos")
            
            # ==================== NUEVA HOJA: REPOSICIÓN NO FULL ====================
            df_no_full = df[df['Reponer Bodega'] > 0].copy()
            
            if not df_no_full.empty:
                reposicion_no_full = df_no_full[['Item ID 1', 'Proveedor', 'SKU',
                                                  'Stock Bodega',  # ⭐ NUEVO: Stock actual en Bodega
                                                  'Ventas FLEX 30d', 'Ventas CENTRO ENVIO 30d',
                                                  'Reponer Bodega',
                                                  'Costo Unitario', 'Ultima Utilidad']].copy()
                
                # Calcular ventas totales no FULL
                reposicion_no_full['Ventas no FULL 30d'] = (reposicion_no_full['Ventas FLEX 30d'] + 
                                                             reposicion_no_full['Ventas CENTRO ENVIO 30d'])
                
                # Calcular costo de reposición
                reposicion_no_full['Costo Reposición'] = (reposicion_no_full['Reponer Bodega'] * 
                                                           reposicion_no_full['Costo Unitario'])
                
                # Reordenar y renombrar columnas
                reposicion_no_full = reposicion_no_full[['Item ID 1', 'Proveedor', 'SKU',
                                                          'Stock Bodega',  # ⭐ NUEVO
                                                          'Ventas no FULL 30d', 'Ventas FLEX 30d', 
                                                          'Ventas CENTRO ENVIO 30d', 'Reponer Bodega',
                                                          'Costo Unitario', 'Ultima Utilidad', 
                                                          'Costo Reposición']]
                
                reposicion_no_full.columns = ['Código Publicación', 'Proveedor', 'SKU',
                                               'Stock Bodega Actual',  # ⭐ NUEVO
                                               'Ventas no FULL 30d', 'Ventas FLEX 30d',
                                               'Ventas Centro Envío 30d', 'Unidades a Reponer',
                                               'Costo Unitario', 'Última Utilidad',
                                               'Costo Reposición']
                
                # Ordenar por proveedor y luego por costo de reposición
                reposicion_no_full = reposicion_no_full.sort_values(['Proveedor', 'Costo Reposición'], 
                                                                     ascending=[True, False])
                
                # Agregar fila de totales
                total_no_full = pd.DataFrame({
                    'Código Publicación': ['TOTAL'],
                    'Proveedor': [''],
                    'SKU': [''],
                    'Stock Bodega Actual': [int(reposicion_no_full['Stock Bodega Actual'].sum())],  # ⭐ NUEVO
                    'Ventas no FULL 30d': [int(reposicion_no_full['Ventas no FULL 30d'].sum())],
                    'Ventas FLEX 30d': [int(reposicion_no_full['Ventas FLEX 30d'].sum())],
                    'Ventas Centro Envío 30d': [int(reposicion_no_full['Ventas Centro Envío 30d'].sum())],
                    'Unidades a Reponer': [int(reposicion_no_full['Unidades a Reponer'].sum())],
                    'Costo Unitario': [''],
                    'Última Utilidad': [reposicion_no_full['Última Utilidad'].mean()],
                    'Costo Reposición': [reposicion_no_full['Costo Reposición'].sum()]
                })
                
                reposicion_no_full = pd.concat([reposicion_no_full, total_no_full], ignore_index=True)
                reposicion_no_full.to_excel(writer, sheet_name='Reposición No FULL', index=False)
                
                print(f"\n✅ Creada hoja 'Reposición No FULL' con {len(df_no_full)} productos")
            
            # ==================== HOJAS POR PROVEEDOR (MEJORADAS) ====================
            for proveedor, group in df.groupby('Proveedor'):
                group_con_reposicion = group[group['Total Reponer'] > 0].copy()
                
                if len(group_con_reposicion) > 0:
                    proveedores_con_reposicion.append(proveedor)
                    
                    sheet_name = str(proveedor)[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
                    
                    # ⭐ TABLA 1: RESUMEN DE REFERENCIA (17 columnas) ⭐
                    orden_compra = group_con_reposicion[['Item ID 1', 'Item ID 2', 'SKU', 'Titulo',
                                        'Ventas FULL 30d', 'Ventas FLEX 30d', 'Ventas CENTRO ENVIO 30d',
                                        'Stock Bodega', 'Reponer Bodega', 
                                        'Stock FULL', 'Reponer FULL', 'Total Reponer',
                                        'Un x Caja Master', 'Cajas Master',
                                        'Costo Unitario', 'Costo Total', 'Ultima Utilidad', '% Rentabilidad']].copy()
                    
                    orden_compra['Ventas no FULL 30d'] = (orden_compra['Ventas FLEX 30d'] + 
                                                          orden_compra['Ventas CENTRO ENVIO 30d'])
                    
                    orden_compra = orden_compra[['Item ID 1', 'Item ID 2', 'SKU', 'Titulo',
                                                'Ventas FULL 30d', 'Ventas no FULL 30d',
                                                'Stock Bodega', 'Reponer Bodega', 
                                                'Stock FULL', 'Reponer FULL', 'Total Reponer',
                                                'Un x Caja Master', 'Cajas Master',
                                                'Costo Unitario', 'Costo Total', 'Ultima Utilidad', '% Rentabilidad']]
                    
                    total_row = pd.DataFrame({
                        'Item ID 1': ['TOTAL'],
                        'Item ID 2': [''],
                        'SKU': [''],
                        'Titulo': [''],
                        'Ventas FULL 30d': [int(group_con_reposicion['Ventas FULL 30d'].sum())],
                        'Ventas no FULL 30d': [int(group_con_reposicion['Ventas FLEX 30d'].sum() + 
                                                   group_con_reposicion['Ventas CENTRO ENVIO 30d'].sum())],
                        'Stock Bodega': [int(group_con_reposicion['Stock Bodega'].max())],
                        'Reponer Bodega': [int(group_con_reposicion['Reponer Bodega'].sum())],
                        'Stock FULL': [int(group_con_reposicion['Stock FULL'].max())],
                        'Reponer FULL': [int(group_con_reposicion['Reponer FULL'].sum())],
                        'Total Reponer': [int(group_con_reposicion['Total Reponer'].sum())],
                        'Un x Caja Master': [''],
                        'Cajas Master': [round(group_con_reposicion['Cajas Master'].sum(), 1)],
                        'Costo Unitario': [''],
                        'Costo Total': [group_con_reposicion['Costo Total'].sum()],
                        'Ultima Utilidad': [''],
                        '% Rentabilidad': [group_con_reposicion['% Rentabilidad'].mean()]
                    })
                    
                    orden_compra = pd.concat([orden_compra, total_row], ignore_index=True)
                    
                    # ⭐ TABLA 2: RESUMEN EJECUTIVO (antes era "Formato Orden de Compra") ⭐
                    tabla_resumen = group_con_reposicion.copy()
                    tabla_resumen['Cajas Master Redondeadas'] = tabla_resumen['Cajas Master'].apply(lambda x: int(round(x, 0)))
                    tabla_resumen['Unidades Totales'] = (tabla_resumen['Cajas Master Redondeadas'] * 
                                                         tabla_resumen['Un x Caja Master']).astype(int)
                    tabla_resumen['Monto Neto'] = (tabla_resumen['Unidades Totales'] * 
                                                   tabla_resumen['Costo Unitario'] / 1.19)
                    tabla_resumen['Monto Bruto'] = (tabla_resumen['Unidades Totales'] * 
                                                    tabla_resumen['Costo Unitario'])
                    tabla_resumen['Unidades Bodega'] = tabla_resumen['Reponer Bodega'].astype(int)
                    tabla_resumen['Unidades FULL'] = tabla_resumen['Reponer FULL'].astype(int)
                    
                    resumen_ejecutivo = tabla_resumen[['Item ID 1', 'SKU', 'Titulo',
                                                       'Cajas Master Redondeadas', 
                                                       'Un x Caja Master',
                                                       'Unidades Totales',
                                                       'Costo Unitario',
                                                       'Monto Neto', 
                                                       'Monto Bruto',
                                                       'Unidades Bodega',
                                                       'Unidades FULL',
                                                       'Ultima Utilidad',
                                                       '% Rentabilidad']].copy()
                    
                    resumen_ejecutivo.columns = ['Código ML', 'SKU', 'Producto',
                                                'Cajas Master', 'Un/Caja', 'Total Unidades',
                                                'Costo Unit.', 'Monto Neto', 'Monto Bruto',
                                                'Unid. Bodega', 'Unid. FULL',
                                                'Última Utilidad', 'Rentabilidad %']
                    
                    total_resumen = pd.DataFrame({
                        'Código ML': ['TOTAL'],
                        'SKU': [''],
                        'Producto': [''],
                        'Cajas Master': [int(resumen_ejecutivo['Cajas Master'].sum())],
                        'Un/Caja': [''],
                        'Total Unidades': [int(resumen_ejecutivo['Total Unidades'].sum())],
                        'Costo Unit.': [''],
                        'Monto Neto': [resumen_ejecutivo['Monto Neto'].sum()],
                        'Monto Bruto': [resumen_ejecutivo['Monto Bruto'].sum()],
                        'Unid. Bodega': [int(resumen_ejecutivo['Unid. Bodega'].sum())],
                        'Unid. FULL': [int(resumen_ejecutivo['Unid. FULL'].sum())],
                        'Última Utilidad': [resumen_ejecutivo['Última Utilidad'].mean()],
                        'Rentabilidad %': [resumen_ejecutivo['Rentabilidad %'].mean()]
                    })
                    
                    resumen_ejecutivo = pd.concat([resumen_ejecutivo, total_resumen], ignore_index=True)
                    
                    # ⭐ TABLA 3: ORDEN DE COMPRA (nueva - para compartir con proveedores) ⭐
                    orden_compra_proveedor = tabla_resumen[['SKU', 'Titulo', 'Unidades Totales',
                                                            'Costo Unitario', 'Monto Neto', 'Monto Bruto']].copy()
                    orden_compra_proveedor.columns = ['SKU', 'Producto', 'Unidades',
                                                      'Valor Unitario', 'Subtotal Neto', 'Subtotal Bruto']
                    
                    total_orden = pd.DataFrame({
                        'SKU': ['TOTAL'],
                        'Producto': [''],
                        'Unidades': [int(orden_compra_proveedor['Unidades'].sum())],
                        'Valor Unitario': [''],
                        'Subtotal Neto': [orden_compra_proveedor['Subtotal Neto'].sum()],
                        'Subtotal Bruto': [orden_compra_proveedor['Subtotal Bruto'].sum()]
                    })
                    
                    orden_compra_proveedor = pd.concat([orden_compra_proveedor, total_orden], ignore_index=True)
                    
                    # ⭐ TABLA 4: BULTOS (nueva) ⭐
                    tabla_bultos = tabla_resumen[['SKU', 'Titulo', 'Unidades Totales', 
                                                  'Un x Caja Master', 
                                                  'Cajas Master Redondeadas']].copy()
                    tabla_bultos.columns = ['SKU', 'Producto', 'Unidades', 'Un/Caja Master', 'Cajas Master']
                    
                    total_bultos = pd.DataFrame({
                        'SKU': ['TOTAL'],
                        'Producto': [''],
                        'Unidades': [int(tabla_bultos['Unidades'].sum())],
                        'Un/Caja Master': [''],
                        'Cajas Master': [int(tabla_bultos['Cajas Master'].sum())]
                    })
                    
                    tabla_bultos = pd.concat([tabla_bultos, total_bultos], ignore_index=True)
                    
                    # ⭐ TABLA 5: VALORES (nueva) ⭐
                    # Calculamos el precio de venta como: Utilidad + Costo
                    tabla_valores = tabla_resumen.copy()
                    tabla_valores['Precio Ultima Venta'] = tabla_valores['Ultima Utilidad'] + tabla_valores['Costo Unitario']
                    
                    valores = tabla_valores[['SKU', 'Titulo', 'Costo Unitario', 'Precio Ultima Venta',
                                            'Ultima Utilidad', '% Rentabilidad']].copy()
                    valores.columns = ['SKU', 'Producto', 'Costo Unitario', 'Precio Última Venta',
                                      'Utilidad', 'Rentabilidad %']
                    
                    total_valores = pd.DataFrame({
                        'SKU': ['PROMEDIO'],
                        'Producto': [''],
                        'Costo Unitario': [valores['Costo Unitario'].mean()],
                        'Precio Última Venta': [valores['Precio Última Venta'].mean()],
                        'Utilidad': [valores['Utilidad'].mean()],
                        'Rentabilidad %': [valores['Rentabilidad %'].mean()]
                    })
                    
                    valores = pd.concat([valores, total_valores], ignore_index=True)
                    
                    # Escribir todas las tablas en la hoja
                    orden_compra.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
                    
                    start_row_resumen = len(orden_compra) + 3
                    resumen_ejecutivo.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_resumen)
                    
                    start_row_orden = start_row_resumen + len(resumen_ejecutivo) + 3
                    orden_compra_proveedor.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_orden)
                    
                    start_row_bultos = start_row_orden + len(orden_compra_proveedor) + 3
                    tabla_bultos.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_bultos)
                    
                    start_row_valores = start_row_bultos + len(tabla_bultos) + 3
                    valores.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row_valores)
            
            # Formatear hoja Detalle
            worksheet_detalle = writer.sheets['Detalle']
            
            for col_idx, column in enumerate(worksheet_detalle.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet_detalle.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar colores por proveedor en hoja Detalle
            for row_idx, row in enumerate(df.itertuples(), start=2):
                proveedor = row.Proveedor
                color = color_map[proveedor]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet_detalle.cell(row=row_idx, column=col)
                    cell.fill = fill
            
            # Aplicar formato de contabilidad a columnas de costos en Detalle
            for row in range(2, len(df) + 2):
                cell_p = worksheet_detalle.cell(row=row, column=16)
                cell_p.number_format = '0.0'
                
                cell_q = worksheet_detalle.cell(row=row, column=17)
                cell_q.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                cell_r = worksheet_detalle.cell(row=row, column=18)
                cell_r.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                cell_s = worksheet_detalle.cell(row=row, column=19)
                cell_s.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                cell_t = worksheet_detalle.cell(row=row, column=20)
                cell_t.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                cell_u = worksheet_detalle.cell(row=row, column=21)
                cell_u.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                cell_v = worksheet_detalle.cell(row=row, column=22)
                cell_v.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Formatear hoja Resumen
            worksheet_resumen = writer.sheets['Resumen']
            
            for col_idx, column in enumerate(worksheet_resumen.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet_resumen.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar colores por proveedor en hoja Resumen
            for row_idx in range(2, len(summary) + 2):
                proveedor = worksheet_resumen.cell(row=row_idx, column=1).value
                if proveedor in color_map:
                    color = color_map[proveedor]
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    for col in range(1, len(summary.columns) + 1):
                        cell = worksheet_resumen.cell(row=row_idx, column=col)
                        cell.fill = fill
            
            # Formato contabilidad en última columna de Resumen
            for row in range(2, len(summary) + 2):
                cell = worksheet_resumen.cell(row=row, column=len(summary.columns))
                cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # ==================== FORMATEAR HOJA REPOSICIÓN FULL ====================
            if not df_full.empty:
                worksheet_full = writer.sheets['Reposición FULL']
                
                fill_header = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                font_total = Font(bold=True)
                fill_total = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                
                # Formatear encabezados (ahora 9 columnas)
                for col in range(1, 10):
                    cell = worksheet_full.cell(row=1, column=col)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatear filas de datos
                for row_idx in range(2, len(df_full) + 2):
                    proveedor = worksheet_full.cell(row=row_idx, column=2).value
                    if proveedor in color_map:
                        color = color_map[proveedor]
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        for col in range(1, 10):
                            cell = worksheet_full.cell(row=row_idx, column=col)
                            cell.fill = fill
                
                # Formatear fila de totales
                total_row_idx = len(df_full) + 2
                for col in range(1, 10):
                    cell = worksheet_full.cell(row=total_row_idx, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato numérico
                for row in range(2, total_row_idx + 1):
                    # Costo Unitario (columna 7)
                    cell_g = worksheet_full.cell(row=row, column=7)
                    cell_g.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Última Utilidad (columna 8)
                    cell_h = worksheet_full.cell(row=row, column=8)
                    cell_h.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Costo Reposición (columna 9)
                    cell_i = worksheet_full.cell(row=row, column=9)
                    cell_i.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Autoajustar columnas
                for col_idx, column in enumerate(worksheet_full.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet_full.column_dimensions[column_letter].width = adjusted_width
            
            # ==================== FORMATEAR HOJA REPOSICIÓN NO FULL ====================
            if not df_no_full.empty:
                worksheet_no_full = writer.sheets['Reposición No FULL']
                
                fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                font_total = Font(bold=True)
                fill_total = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                
                # Formatear encabezados (ahora 11 columnas)
                for col in range(1, 12):
                    cell = worksheet_no_full.cell(row=1, column=col)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatear filas de datos
                for row_idx in range(2, len(df_no_full) + 2):
                    proveedor = worksheet_no_full.cell(row=row_idx, column=2).value
                    if proveedor in color_map:
                        color = color_map[proveedor]
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        for col in range(1, 12):
                            cell = worksheet_no_full.cell(row=row_idx, column=col)
                            cell.fill = fill
                
                # Formatear fila de totales
                total_row_idx = len(df_no_full) + 2
                for col in range(1, 12):
                    cell = worksheet_no_full.cell(row=total_row_idx, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato numérico
                for row in range(2, total_row_idx + 1):
                    # Costo Unitario (columna 9)
                    cell_i = worksheet_no_full.cell(row=row, column=9)
                    cell_i.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Última Utilidad (columna 10)
                    cell_j = worksheet_no_full.cell(row=row, column=10)
                    cell_j.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Costo Reposición (columna 11)
                    cell_k = worksheet_no_full.cell(row=row, column=11)
                    cell_k.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Autoajustar columnas
                for col_idx, column in enumerate(worksheet_no_full.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet_no_full.column_dimensions[column_letter].width = adjusted_width
            
            # ⭐ FORMATEAR HOJAS DE PROVEEDORES CON REPOSICIÓN ⭐
            for proveedor in proveedores_con_reposicion:
                group = df[(df['Proveedor'] == proveedor) & (df['Total Reponer'] > 0)]
                
                sheet_name = str(proveedor)[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
                worksheet = writer.sheets[sheet_name]
                
                color = color_map[proveedor]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                fill_header_orden = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                font_total = Font(bold=True)
                fill_total = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                
                # ===== TABLA 1: RESUMEN DE REFERENCIA =====
                # ⭐ AHORA SON 17 COLUMNAS (removidas FLEX y CENTRO ENVIO individuales)
                for col in range(1, 18):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                total_rows_table1 = len(group) + 1
                fill_perdida = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                font_perdida = Font(bold=True, color='FFFFFF')
                
                for row_idx in range(2, total_rows_table1 + 1):
                    # Primero aplicar color normal del proveedor
                    for col in range(1, 18):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                    
                    # Verificar si la utilidad es negativa (columna 16)
                    utilidad_cell = worksheet.cell(row=row_idx, column=16)
                    if utilidad_cell.value is not None and isinstance(utilidad_cell.value, (int, float)):
                        if utilidad_cell.value < 0:
                            # Aplicar fondo rojo a toda la fila si hay pérdida
                            for col in range(1, 18):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = fill_perdida
                                cell.font = font_perdida
                
                # Formatear fila de totales de tabla 1
                for col in range(1, 18):
                    cell = worksheet.cell(row=total_rows_table1 + 1, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato de contabilidad y numérico de tabla 1
                for row in range(2, total_rows_table1 + 2):
                    # Cajas Master (columna 13)
                    cell_m = worksheet.cell(row=row, column=13)
                    cell_m.number_format = '0.0'
                    
                    # Costo Unitario (columna 14)
                    cell_n = worksheet.cell(row=row, column=14)
                    cell_n.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Costo Total (columna 15)
                    cell_o = worksheet.cell(row=row, column=15)
                    cell_o.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Ultima Utilidad (columna 16)
                    cell_p = worksheet.cell(row=row, column=16)
                    cell_p.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # % Rentabilidad (columna 17)
                    cell_q = worksheet.cell(row=row, column=17)
                    cell_q.number_format = '0.00"%"'
                
                # ===== TABLA 2: FORMATO ORDEN DE COMPRA =====
                start_row_table2 = total_rows_table1 + 4
                
                # Agregar título para la tabla 2
                title_row = start_row_table2 - 1
                worksheet.cell(row=title_row, column=1).value = "FORMATO ORDEN DE COMPRA"
                title_cell = worksheet.cell(row=title_row, column=1)
                title_cell.font = Font(bold=True, size=14, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # ⭐ Ahora son 13 columnas (agregadas Unid. Bodega y Unid. FULL)
                worksheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=13)
                
                # Formatear encabezados de la tabla 2 (13 columnas)
                for col in range(1, 14):
                    cell = worksheet.cell(row=start_row_table2, column=col)
                    cell.fill = fill_header_orden
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar color a filas de datos de tabla 2
                total_rows_table2 = len(group) + 1
                for row_idx in range(start_row_table2 + 1, start_row_table2 + total_rows_table2):
                    # Primero aplicar color normal del proveedor
                    for col in range(1, 14):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                    
                    # Verificar si la utilidad es negativa (columna 12)
                    utilidad_cell = worksheet.cell(row=row_idx, column=12)
                    if utilidad_cell.value is not None and isinstance(utilidad_cell.value, (int, float)):
                        if utilidad_cell.value < 0:
                            # Aplicar fondo rojo a toda la fila si hay pérdida
                            for col in range(1, 14):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = fill_perdida
                                cell.font = font_perdida
                
                # Formatear fila de totales de tabla 2
                last_row_table2 = start_row_table2 + total_rows_table2
                for col in range(1, 14):
                    cell = worksheet.cell(row=last_row_table2, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato de contabilidad y numérico de tabla 2
                for row in range(start_row_table2 + 1, last_row_table2 + 1):
                    # Costo Unitario (columna 7)
                    cell_g = worksheet.cell(row=row, column=7)
                    cell_g.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Monto Neto (columna 8)
                    cell_h = worksheet.cell(row=row, column=8)
                    cell_h.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Monto Bruto (columna 9)
                    cell_i = worksheet.cell(row=row, column=9)
                    cell_i.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Unid. Bodega (columna 10) - formato entero
                    # Unid. FULL (columna 11) - formato entero
                    
                    # Última Utilidad (columna 12)
                    cell_l = worksheet.cell(row=row, column=12)
                    cell_l.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Rentabilidad % (columna 13)
                    cell_m = worksheet.cell(row=row, column=13)
                    cell_m.number_format = '0.00"%"'
                
                # ===== TABLA 3: ORDEN DE COMPRA (para proveedores) =====
                start_row_table3 = last_row_table2 + 3
                
                # Agregar título para la tabla 3
                title_row_table3 = start_row_table3 - 1
                worksheet.cell(row=title_row_table3, column=1).value = "ORDEN DE COMPRA - PROVEEDOR"
                title_cell_table3 = worksheet.cell(row=title_row_table3, column=1)
                title_cell_table3.font = Font(bold=True, size=14, color='FFFFFF')
                title_cell_table3.fill = PatternFill(start_color='9966FF', end_color='9966FF', fill_type='solid')
                title_cell_table3.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.merge_cells(start_row=title_row_table3, start_column=1, end_row=title_row_table3, end_column=6)
                
                # Colores para tabla 3 (morado)
                fill_header_table3 = PatternFill(start_color='9966FF', end_color='9966FF', fill_type='solid')
                
                # Formatear encabezados de la tabla 3 (6 columnas)
                for col in range(1, 7):
                    cell = worksheet.cell(row=start_row_table3, column=col)
                    cell.fill = fill_header_table3
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar color a filas de datos de tabla 3
                total_rows_table3 = len(group) + 1
                for row_idx in range(start_row_table3 + 1, start_row_table3 + total_rows_table3):
                    for col in range(1, 7):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                
                # Formatear fila de totales de tabla 3
                last_row_table3 = start_row_table3 + total_rows_table3
                for col in range(1, 7):
                    cell = worksheet.cell(row=last_row_table3, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato de contabilidad y numérico de tabla 3
                for row in range(start_row_table3 + 1, last_row_table3 + 1):
                    # Valor Unitario (columna 4)
                    cell_d = worksheet.cell(row=row, column=4)
                    cell_d.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Subtotal Neto (columna 5)
                    cell_e = worksheet.cell(row=row, column=5)
                    cell_e.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Subtotal Bruto (columna 6)
                    cell_f = worksheet.cell(row=row, column=6)
                    cell_f.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # ===== TABLA 4: BULTOS (logística) =====
                start_row_table4 = last_row_table3 + 3
                
                # Agregar título para la tabla 4
                title_row_table4 = start_row_table4 - 1
                worksheet.cell(row=title_row_table4, column=1).value = "LOGÍSTICA - EMPAQUE"
                title_cell_table4 = worksheet.cell(row=title_row_table4, column=1)
                title_cell_table4.font = Font(bold=True, size=14, color='FFFFFF')
                title_cell_table4.fill = PatternFill(start_color='9966FF', end_color='9966FF', fill_type='solid')
                title_cell_table4.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.merge_cells(start_row=title_row_table4, start_column=1, end_row=title_row_table4, end_column=5)
                
                # Colores para tabla 4 (morado)
                fill_header_table4 = PatternFill(start_color='9966FF', end_color='9966FF', fill_type='solid')
                
                # Formatear encabezados de la tabla 4 (5 columnas - agregamos Producto)
                for col in range(1, 6):
                    cell = worksheet.cell(row=start_row_table4, column=col)
                    cell.fill = fill_header_table4
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar color a filas de datos de tabla 4
                total_rows_table4 = len(group) + 1
                for row_idx in range(start_row_table4 + 1, start_row_table4 + total_rows_table4):
                    for col in range(1, 6):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                
                # Formatear fila de totales de tabla 4
                last_row_table4 = start_row_table4 + total_rows_table4
                for col in range(1, 6):
                    cell = worksheet.cell(row=last_row_table4, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # ===== TABLA 5: VALORES (rentabilidad) =====
                start_row_table5 = last_row_table4 + 3
                
                # Agregar título para la tabla 5
                title_row_table5 = start_row_table5 - 1
                worksheet.cell(row=title_row_table5, column=1).value = "ANÁLISIS DE RENTABILIDAD"
                title_cell_table5 = worksheet.cell(row=title_row_table5, column=1)
                title_cell_table5.font = Font(bold=True, size=14, color='FFFFFF')
                title_cell_table5.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                title_cell_table5.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.merge_cells(start_row=title_row_table5, start_column=1, end_row=title_row_table5, end_column=6)
                
                # Colores para tabla 5 (azul)
                fill_header_table5 = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                
                # Formatear encabezados de la tabla 5 (6 columnas - agregamos Producto)
                for col in range(1, 7):
                    cell = worksheet.cell(row=start_row_table5, column=col)
                    cell.fill = fill_header_table5
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar color a filas de datos de tabla 5
                total_rows_table5 = len(group) + 1
                fill_perdida = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                font_perdida = Font(bold=True, color='FFFFFF')
                
                for row_idx in range(start_row_table5 + 1, start_row_table5 + total_rows_table5):
                    # Primero aplicar color normal del proveedor
                    for col in range(1, 7):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                    
                    # Verificar si la utilidad es negativa (columna 5)
                    utilidad_cell = worksheet.cell(row=row_idx, column=5)
                    if utilidad_cell.value is not None and isinstance(utilidad_cell.value, (int, float)):
                        if utilidad_cell.value < 0:
                            # Aplicar fondo rojo a toda la fila si hay pérdida
                            for col in range(1, 7):
                                cell = worksheet.cell(row=row_idx, column=col)
                                cell.fill = fill_perdida
                                cell.font = font_perdida
                
                # Formatear fila de totales de tabla 5 (PROMEDIO)
                last_row_table5 = start_row_table5 + total_rows_table5
                for col in range(1, 7):
                    cell = worksheet.cell(row=last_row_table5, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato de contabilidad y numérico de tabla 5
                for row in range(start_row_table5 + 1, last_row_table5 + 1):
                    # Costo Unitario (columna 3 - antes era 2)
                    cell_c = worksheet.cell(row=row, column=3)
                    cell_c.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Precio Última Venta (columna 4 - antes era 3)
                    cell_d = worksheet.cell(row=row, column=4)
                    cell_d.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Utilidad (columna 5 - antes era 4)
                    cell_e = worksheet.cell(row=row, column=5)
                    cell_e.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Rentabilidad % (columna 6 - antes era 5) - 1 decimal
                    cell_f = worksheet.cell(row=row, column=6)
                    cell_f.number_format = '0.0"%"'
                
                # Autoajustar columnas
                for col_idx, column in enumerate(worksheet.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # ==================== NUEVA HOJA: PRODUCTOS SIN STOCK ====================
            print("\n📦 Creando hoja de productos sin stock...")
            
            # Filtrar productos sin stock (0 en bodega Y 0 en FULL total)
            df_sin_stock = df[(df['Stock Bodega'] == 0) & (df['Stock FULL Total'] == 0)].copy()
            
            if not df_sin_stock.empty:
                # Preparar datos para la hoja
                sin_stock_data = df_sin_stock[['Item ID 1', 'SKU', 'Titulo', 'Proveedor',
                                                'Precio Venta', 'Ultima Utilidad', 
                                                'Fecha Ultima Venta']].copy()
                
                # Renombrar columnas
                sin_stock_data.columns = ['Código Publicación', 'SKU', 'Producto', 'Proveedor',
                                          'Precio Última Venta', 'Utilidad Última Venta', 
                                          'Fecha Última Venta']
                
                # Convertir fecha a formato legible, manejando valores nulos
                sin_stock_data['Fecha Última Venta'] = pd.to_datetime(
                    sin_stock_data['Fecha Última Venta'], errors='coerce'
                )
                
                # Formatear solo las fechas válidas
                sin_stock_data['Fecha Última Venta'] = sin_stock_data['Fecha Última Venta'].apply(
                    lambda x: x.strftime('%Y-%m-%d %H:%M') if pd.notna(x) else 'Sin ventas recientes'
                )
                
                # Ordenar por proveedor y luego por utilidad descendente
                sin_stock_data = sin_stock_data.sort_values(['Proveedor', 'Utilidad Última Venta'], 
                                                             ascending=[True, False])
                
                # Escribir a Excel
                sin_stock_data.to_excel(writer, sheet_name='Sin Stock', index=False)
                
                # Formatear la hoja
                worksheet_sin_stock = writer.sheets['Sin Stock']
                
                fill_header = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                
                # Formatear encabezados
                for col in range(1, 8):
                    cell = worksheet_sin_stock.cell(row=1, column=col)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar colores por proveedor
                for row_idx in range(2, len(sin_stock_data) + 2):
                    proveedor = worksheet_sin_stock.cell(row=row_idx, column=4).value
                    if proveedor in color_map:
                        color = color_map[proveedor]
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        for col in range(1, 8):
                            cell = worksheet_sin_stock.cell(row=row_idx, column=col)
                            cell.fill = fill
                
                # Formato numérico
                for row in range(2, len(sin_stock_data) + 2):
                    # Precio Última Venta (columna 5)
                    cell_e = worksheet_sin_stock.cell(row=row, column=5)
                    cell_e.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Utilidad Última Venta (columna 6)
                    cell_f = worksheet_sin_stock.cell(row=row, column=6)
                    cell_f.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Autoajustar columnas
                for col_idx, column in enumerate(worksheet_sin_stock.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet_sin_stock.column_dimensions[column_letter].width = adjusted_width
                
                print(f"✅ Creada hoja 'Sin Stock' con {len(sin_stock_data)} productos")
            else:
                print("✅ No hay productos sin stock")
        
        print(f"\n✅ Reporte guardado: {filename}")
        print(f"📋 Hojas creadas:")
        print(f"   - Detalle")
        print(f"   - Resumen")
        if not df_full.empty:
            print(f"   - Reposición FULL ({len(df_full)} productos)")
        if not df_no_full.empty:
            print(f"   - Reposición No FULL ({len(df_no_full)} productos)")
        
        # Contar productos sin stock
        df_sin_stock_count = df[(df['Stock Bodega'] == 0) & (df['Stock FULL Total'] == 0)]
        if not df_sin_stock_count.empty:
            print(f"   - Sin Stock ({len(df_sin_stock_count)} productos)")
        
        print(f"   - {len(proveedores_con_reposicion)} hojas de proveedores")
        print("\n🎯 MEJORAS IMPLEMENTADAS:")
        print("   ✅ Nueva hoja 'Reposición FULL' con stock actual visible")
        print("   ✅ Nueva hoja 'Reposición No FULL' con stock actual y desglose FLEX/Mercado Envíos")
        print("   ✅ Nueva hoja 'Sin Stock' con fecha y datos de última venta")
        print("   ✅ Hojas de proveedores reorganizadas en 5 tablas:")
        print("      1. 📘 Resumen de Referencia (17 cols)")
        print("      2. 💚 Resumen Ejecutivo (13 cols) - Info interna completa")
        print("      3. 🧡 Orden de Compra (6 cols) - Para compartir con proveedores")
        print("      4. 💜 Bultos (5 cols) - Información logística con nombre de producto")
        print("      5. 📘 Valores (6 cols) - Análisis de rentabilidad con nombre de producto")
        print("   🔴 Productos con pérdidas resaltados en rojo")
    
    def generate_inventory_report(self, df):
        """Genera reporte de valorización de inventario actual
        MODIFICADO: Considera unidades en tránsito en valorización
        """
        if df.empty:
            return
        
        print("\n" + "="*80)
        print("GENERANDO REPORTE DE VALORIZACIÓN DE INVENTARIO")
        print("="*80)
        
        # Calcular valorización
        df_inventario = df.copy()
        df_inventario['Valor Stock Bodega'] = df_inventario['Stock Bodega'] * df_inventario['Costo Unitario']
        # MODIFICADO: Usar Stock FULL Total que incluye tránsito
        df_inventario['Valor Stock FULL'] = df_inventario['Stock FULL Total'] * df_inventario['Costo Unitario']
        df_inventario['Valor Total Stock'] = df_inventario['Valor Stock Bodega'] + df_inventario['Valor Stock FULL']
        df_inventario['Stock Total'] = df_inventario['Stock Bodega'] + df_inventario['Stock FULL Total']
        
        # Resumen por proveedor
        summary_inventario = df_inventario.groupby('Proveedor').agg({
            'SKU': 'count',
            'Stock Bodega': 'sum',
            'Valor Stock Bodega': 'sum',
            'Stock FULL Total': 'sum',  # MODIFICADO: Usar total con tránsito
            'En Tránsito FULL': 'sum',  # NUEVO
            'Valor Stock FULL': 'sum',
            'Stock Total': 'sum',
            'Valor Total Stock': 'sum'
        }).reset_index()
        
        summary_inventario.columns = [
            'Proveedor', 'Items', 'Unid Bodega', 'Valor Bodega',
            'Unid FULL Total', 'En Tránsito', 'Valor FULL', 'Total Unidades', 'Valor Total'  # MODIFICADO
        ]
        
        # Imprimir resumen en consola
        print("\nRESUMEN DE INVENTARIO POR PROVEEDOR:")
        print(summary_inventario.to_string(index=False))
        print(f"\nTotal unidades en inventario: {int(df_inventario['Stock Total'].sum())}")
        print(f"  - Stock FULL Total: {int(df_inventario['Stock FULL Total'].sum())} (incluye {int(df_inventario['En Tránsito FULL'].sum())} en tránsito)")
        print(f"  - Stock Bodega: {int(df_inventario['Stock Bodega'].sum())}")
        print(f"Valor total del inventario: ${df_inventario['Valor Total Stock'].sum():,.0f}")
        
        # Colores por proveedor (mismos que reporte de reposición)
        colores_proveedores = [
            'FFB6C1', 'B0E0E6', '98FB98', 'FFE4B5', 'DDA0DD',
            'F0E68C', 'FFA07A', '87CEEB', 'F5DEB3', 'D8BFD8',
        ]
        
        proveedores_unicos = df_inventario['Proveedor'].unique()
        color_map = {proveedor: colores_proveedores[i % len(colores_proveedores)] 
                    for i, proveedor in enumerate(proveedores_unicos)}
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"valorizacion_inventario_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Hoja de Detalle - MODIFICADO: Agregar columnas de tránsito
            df_detalle_inv = df_inventario[[
                'Item ID 1', 'Item ID 2', 'SKU', 'Titulo', 'Proveedor',
                'Stock Bodega', 'Costo Unitario', 'Valor Stock Bodega',
                'Stock FULL', 'En Tránsito FULL', 'Stock FULL Total', 'Valor Stock FULL',  # MODIFICADO
                'Stock Total', 'Valor Total Stock'
            ]].copy()
            
            df_detalle_inv.to_excel(writer, sheet_name='Detalle Inventario', index=False)
            
            # Hoja de Resumen
            summary_inventario.to_excel(writer, sheet_name='Resumen por Proveedor', index=False)
            
            # Hojas por proveedor
            for proveedor in proveedores_unicos:
                group = df_inventario[df_inventario['Proveedor'] == proveedor].copy()
                
                sheet_name = str(proveedor)[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
                
                # Seleccionar columnas para la hoja del proveedor - MODIFICADO
                inventario_proveedor = group[[
                    'Item ID 1', 'Item ID 2', 'SKU', 'Titulo',
                    'Stock Bodega', 'Costo Unitario', 'Valor Stock Bodega',
                    'Stock FULL', 'Valor Stock FULL',
                    'Stock Total', 'Valor Total Stock'
                ]].copy()
                
                # Agregar fila de totales
                total_row = pd.DataFrame({
                    'Item ID 1': ['TOTAL'],
                    'Item ID 2': [''],
                    'SKU': [''],
                    'Titulo': [''],
                    'Stock Bodega': [int(group['Stock Bodega'].sum())],
                    'Costo Unitario': [''],
                    'Valor Stock Bodega': [group['Valor Stock Bodega'].sum()],
                    'Stock FULL': [int(group['Stock FULL'].sum())],
                    'Valor Stock FULL': [group['Valor Stock FULL'].sum()],
                    'Stock Total': [int(group['Stock Total'].sum())],
                    'Valor Total Stock': [group['Valor Total Stock'].sum()]
                })
                
                inventario_proveedor = pd.concat([inventario_proveedor, total_row], ignore_index=True)
                inventario_proveedor.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Formatear hojas
            workbook = writer.book
            
            # Formatear hoja Detalle Inventario
            worksheet_detalle = writer.sheets['Detalle Inventario']
            
            for col_idx, column in enumerate(worksheet_detalle.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet_detalle.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar colores por proveedor en hoja Detalle
            for row_idx, row in enumerate(df_detalle_inv.itertuples(), start=2):
                proveedor = row.Proveedor
                color = color_map[proveedor]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
                for col in range(1, len(df_detalle_inv.columns) + 1):
                    cell = worksheet_detalle.cell(row=row_idx, column=col)
                    cell.fill = fill
            
            # Aplicar formato de contabilidad a columnas de valores en Detalle
            for row in range(2, len(df_detalle_inv) + 2):
                # Costo Unitario (columna 7)
                cell_g = worksheet_detalle.cell(row=row, column=7)
                cell_g.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Valor Stock Bodega (columna 8)
                cell_h = worksheet_detalle.cell(row=row, column=8)
                cell_h.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Valor Stock FULL (columna 10)
                cell_j = worksheet_detalle.cell(row=row, column=10)
                cell_j.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Valor Total Stock (columna 12)
                cell_l = worksheet_detalle.cell(row=row, column=12)
                cell_l.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Formatear hoja Resumen
            worksheet_resumen = writer.sheets['Resumen por Proveedor']
            
            for col_idx, column in enumerate(worksheet_resumen.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet_resumen.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar colores por proveedor en hoja Resumen
            for row_idx in range(2, len(summary_inventario) + 2):
                proveedor = worksheet_resumen.cell(row=row_idx, column=1).value
                if proveedor in color_map:
                    color = color_map[proveedor]
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    for col in range(1, len(summary_inventario.columns) + 1):
                        cell = worksheet_resumen.cell(row=row_idx, column=col)
                        cell.fill = fill
            
            # Formato contabilidad en columnas de valores del Resumen
            for row in range(2, len(summary_inventario) + 2):
                # Valor Bodega (columna 4)
                cell_d = worksheet_resumen.cell(row=row, column=4)
                cell_d.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Valor FULL (columna 6)
                cell_f = worksheet_resumen.cell(row=row, column=6)
                cell_f.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Valor Total (columna 8)
                cell_h = worksheet_resumen.cell(row=row, column=8)
                cell_h.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # Formatear hojas de proveedores
            for proveedor in proveedores_unicos:
                group = df_inventario[df_inventario['Proveedor'] == proveedor]
                
                sheet_name = str(proveedor)[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
                worksheet = writer.sheets[sheet_name]
                
                color = color_map[proveedor]
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                font_header = Font(bold=True, color='FFFFFF')
                font_total = Font(bold=True)
                fill_total = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                
                # Formatear encabezados
                for col in range(1, 12):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar color a filas de datos
                total_rows = len(group) + 1
                for row_idx in range(2, total_rows + 1):
                    for col in range(1, 12):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = fill
                
                # Formatear fila de totales
                for col in range(1, 12):
                    cell = worksheet.cell(row=total_rows + 1, column=col)
                    cell.fill = fill_total
                    cell.font = font_total
                
                # Formato de contabilidad y numérico
                for row in range(2, total_rows + 2):
                    # Costo Unitario (columna 6)
                    cell_f = worksheet.cell(row=row, column=6)
                    cell_f.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Valor Stock Bodega (columna 7)
                    cell_g = worksheet.cell(row=row, column=7)
                    cell_g.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Valor Stock FULL (columna 9)
                    cell_i = worksheet.cell(row=row, column=9)
                    cell_i.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                    
                    # Valor Total Stock (columna 11)
                    cell_k = worksheet.cell(row=row, column=11)
                    cell_k.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                
                # Autoajustar columnas
                for col_idx, column in enumerate(worksheet.columns, start=1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    
                    for cell in column:
                        try:
                            if hasattr(cell, 'value') and cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"\nReporte de inventario guardado: {filename}")
        print(f"Hojas creadas: Detalle Inventario, Resumen por Proveedor + {len(proveedores_unicos)} hojas de proveedores")

    def analyze_monthly_profits(self, test_mode=False, test_limit=50, months_back=6, only_active=False):
        """Analiza utilidades por mes"""
        print("\n" + "="*80)
        if test_mode:
            print(f"ANÁLISIS DE UTILIDADES MENSUALES - MODO PRUEBA ({test_limit} items)")
        else:
            print("ANÁLISIS DE UTILIDADES MENSUALES")
        print(f"Analizando últimos {months_back} meses")
        if only_active:
            print("⚡ FILTRO ACTIVADO - Solo productos activos")
        print("="*80)
        
        # Obtener todas las publicaciones
        items = self.get_all_items(only_active=only_active)
        if not items:
            return None
        
        if test_mode:
            items = items[:test_limit]
            print(f"\n✓ MODO PRUEBA: Analizando solo {len(items)} publicaciones")
        
        # Cargar Excel para obtener SKU, Proveedor y Costo
        excel_data = self._load_excel()
        if excel_data is None:
            print("⚠ No se pudo cargar Excel, continuando sin SKU/Proveedor/Costo")
            excel_dict = {}
        else:
            excel_dict = self._create_excel_dict(excel_data)
        
        # Generar lista de meses a analizar
        today = datetime.now()
        months_to_analyze = []
        
        for i in range(months_back):
            month_date = today - timedelta(days=30*i)
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # Calcular el último día del mes
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)
            
            month_end = month_end.replace(hour=23, minute=59, second=59)
            
            months_to_analyze.append({
                'name': month_start.strftime('%Y-%m'),
                'display_name': month_start.strftime('%B %Y'),
                'start': month_start,
                'end': month_end if month_date.month != today.month else today
            })
        
        months_to_analyze.reverse()  # Ordenar de más antiguo a más reciente
        
        print(f"\nMeses a analizar:")
        for month in months_to_analyze:
            print(f"  - {month['display_name']}")
        
        # Diccionario para almacenar datos por mes
        monthly_data = {month['name']: [] for month in months_to_analyze}
        
        # Analizar cada producto
        print("\nAnalizando ventas por producto...")
        for idx, item_id in enumerate(items, 1):
            item_details = self.get_item_details(item_id)
            if not item_details:
                continue
            
            title = item_details.get('title', 'N/A')
            print(f"\n[{idx}/{len(items)}] {item_id} - {title[:50]}...")
            
            # Obtener info del Excel
            excel_info = excel_dict.get(item_id, {})
            sku = excel_info.get('SKU', 'N/A')
            proveedor = excel_info.get('Proveedor', 'N/A')
            costo_unitario = excel_info.get('Costo', 0)
            
            # Analizar cada mes
            for month in months_to_analyze:
                date_from = month['start'].strftime('%Y-%m-%dT%H:%M:%S.000-04:00')
                date_to = month['end'].strftime('%Y-%m-%dT%H:%M:%S.000-04:00')
                
                url = f"{self.base_url}/orders/search"
                headers = {'Authorization': f'Bearer {self.access_token}'}
                params = {
                    'seller': self.user_id,
                    'item': item_id,
                    'order.date_created.from': date_from,
                    'order.date_created.to': date_to
                }
                
                try:
                    response = requests.get(url, headers=headers, params=params)
                    response.raise_for_status()
                    data = response.json()
                    
                    results = data.get('results', [])
                    if not results:
                        continue
                    
                    print(f"  {month['display_name']}: {len(results)} órdenes")
                    
                    # Procesar cada orden
                    for order_data in results:
                        try:
                            payments = order_data.get('payments')
                            if not payments:
                                continue
                            
                            order_id = payments[0]['order_id']
                            order_detail = self._get_order_detail(order_id)
                            
                            if not order_detail or order_detail.get('status') != 'paid':
                                continue
                            
                            quantity = self._get_item_quantity(order_detail, item_id)
                            if quantity == 0:
                                continue
                            
                            # Obtener información de la orden
                            order_items = order_detail.get('order_items', [])
                            item_info = None
                            for item in order_items:
                                if item.get('item', {}).get('id') == item_id:
                                    item_info = item
                                    break
                            
                            if not item_info:
                                continue
                            
                            precio_unitario = item_info.get('unit_price', 0)
                            comision = item_info.get('sale_fee', 0)
                            
                            # Obtener costo de envío
                            costo_envio = 0
                            shipping_id = order_detail.get('shipping', {}).get('id')
                            if shipping_id:
                                shipping_costs = self._get_costs_info(shipping_id)
                                if shipping_costs and shipping_costs.get('senders'):
                                    costo_envio = shipping_costs['senders'][0].get('cost', 0)
                            
                            # Calcular métricas
                            facturacion = precio_unitario * quantity
                            costo_total = costo_unitario * quantity
                            comision_total = comision * quantity
                            costo_envio_total = costo_envio
                            
                            ingreso_neto = facturacion - comision_total - costo_envio_total
                            utilidad = ingreso_neto - costo_total
                            rentabilidad = (utilidad / facturacion * 100) if facturacion > 0 else 0
                            
                            monthly_data[month['name']].append({
                                'Item ID': item_id,
                                'SKU': sku,
                                'Título': title,
                                'Proveedor': proveedor,
                                'Cantidad': quantity,
                                'Precio Unitario': precio_unitario,
                                'Costo Unitario': costo_unitario,
                                'Facturación': facturacion,
                                'Comisión': comision_total,
                                'Costo Envío': costo_envio_total,
                                'Costo Total': costo_total,
                                'Ingreso Neto': ingreso_neto,
                                'Utilidad': utilidad,
                                'Rentabilidad %': rentabilidad
                            })
                            
                        except Exception as e:
                            continue
                    
                except Exception as e:
                    print(f"  Error en {month['display_name']}: {e}")
                    continue
        
        return monthly_data, months_to_analyze

    def generate_monthly_profits_report(self, monthly_data, months_info):
        """Genera reporte Excel de utilidades mensuales"""
        if not monthly_data:
            print("No hay datos para generar reporte")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"utilidades_mensuales_{timestamp}.xlsx"    
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Estilos
        font_header = Font(bold=True, size=11, color='FFFFFF')
        fill_header = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        fill_total = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        font_total = Font(bold=True, size=11)
        alignment_center = Alignment(horizontal='center', vertical='center')
        
        # Crear una hoja por mes
        for month in months_info:
            month_name = month['name']
            month_display = month['display_name']
            data = monthly_data.get(month_name, [])
            
            if not data:
                continue
            
            # Crear DataFrame
            df = pd.DataFrame(data)
            
            # Calcular totales
            total_facturacion = df['Facturación'].sum()
            total_comision = df['Comisión'].sum()
            total_costo_envio = df['Costo Envío'].sum()
            total_costo = df['Costo Total'].sum()
            total_ingreso_neto = df['Ingreso Neto'].sum()
            total_utilidad = df['Utilidad'].sum()
            rentabilidad_promedio = (total_utilidad / total_facturacion * 100) if total_facturacion > 0 else 0
            
            ws = wb.create_sheet(title=month_name)
            
            # TÍTULO Y RESUMEN
            ws['A1'] = f'UTILIDADES - {month_display.upper()}'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:G1')
            
            ws['A3'] = 'RESUMEN DEL MES'
            ws['A3'].font = font_header
            ws['A3'].fill = fill_header
            ws.merge_cells('A3:B3')
            
            resumen_data = [
                ['Facturación Total:', total_facturacion],
                ['Comisiones:', total_comision],
                ['Costos de Envío:', total_costo_envio],
                ['Costos de Productos:', total_costo],
                ['Ingreso Neto:', total_ingreso_neto],
                ['Utilidad Total:', total_utilidad],
                ['Rentabilidad Promedio:', f'{rentabilidad_promedio:.2f}%']
            ]
            
            for idx, (label, value) in enumerate(resumen_data, start=4):
                ws[f'A{idx}'] = label
                ws[f'A{idx}'].font = Font(bold=True)
                if isinstance(value, str):
                    ws[f'B{idx}'] = value
                else:
                    ws[f'B{idx}'] = value
                    ws[f'B{idx}'].number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
            
            # TOP 10 PRODUCTOS POR UTILIDAD
            start_row = 13
            ws[f'A{start_row}'] = 'TOP 10 PRODUCTOS POR UTILIDAD'
            ws[f'A{start_row}'].font = font_header
            ws[f'A{start_row}'].fill = fill_header
            ws.merge_cells(f'A{start_row}:F{start_row}')
            
            df_productos = df.groupby(['Item ID', 'Título', 'SKU', 'Proveedor']).agg({
                'Cantidad': 'sum',
                'Facturación': 'sum',
                'Utilidad': 'sum',
                'Rentabilidad %': 'mean'
            }).reset_index()
            
            df_productos = df_productos.sort_values('Utilidad', ascending=False).head(10)
            
            headers_productos = ['Item ID', 'Título', 'SKU', 'Proveedor', 'Cantidad', 'Facturación', 'Utilidad', 'Rentabilidad %']
            for col_idx, header in enumerate(headers_productos, start=1):
                cell = ws.cell(row=start_row+1, column=col_idx)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = alignment_center
            
            for row_idx, (_, row) in enumerate(df_productos.iterrows(), start=start_row+2):
                ws.cell(row=row_idx, column=1, value=row['Item ID'])
                ws.cell(row=row_idx, column=2, value=row['Título'][:50])
                ws.cell(row=row_idx, column=3, value=row['SKU'])
                ws.cell(row=row_idx, column=4, value=row['Proveedor'])
                ws.cell(row=row_idx, column=5, value=row['Cantidad'])
                ws.cell(row=row_idx, column=6, value=row['Facturación'])
                ws.cell(row=row_idx, column=6).number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                ws.cell(row=row_idx, column=7, value=row['Utilidad'])
                ws.cell(row=row_idx, column=7).number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                ws.cell(row=row_idx, column=8, value=row['Rentabilidad %'])
                ws.cell(row=row_idx, column=8).number_format = '0.00"%"'
            
            # RESUMEN POR PROVEEDOR
            start_row_prov = start_row + len(df_productos) + 4
            ws[f'A{start_row_prov}'] = 'RESUMEN POR PROVEEDOR'
            ws[f'A{start_row_prov}'].font = font_header
            ws[f'A{start_row_prov}'].fill = fill_header
            ws.merge_cells(f'A{start_row_prov}:E{start_row_prov}')
            
            df_proveedores = df.groupby('Proveedor').agg({
                'Cantidad': 'sum',
                'Facturación': 'sum',
                'Utilidad': 'sum',
                'Rentabilidad %': 'mean'
            }).reset_index()
            
            df_proveedores = df_proveedores.sort_values('Utilidad', ascending=False)
            
            headers_prov = ['Proveedor', 'Cantidad Vendida', 'Facturación', 'Utilidad', 'Rentabilidad Promedio']
            for col_idx, header in enumerate(headers_prov, start=1):
                cell = ws.cell(row=start_row_prov+1, column=col_idx)
                cell.value = header
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = alignment_center
            
            for row_idx, (_, row) in enumerate(df_proveedores.iterrows(), start=start_row_prov+2):
                ws.cell(row=row_idx, column=1, value=row['Proveedor'])
                ws.cell(row=row_idx, column=2, value=row['Cantidad'])
                ws.cell(row=row_idx, column=3, value=row['Facturación'])
                ws.cell(row=row_idx, column=3).number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                ws.cell(row=row_idx, column=4, value=row['Utilidad'])
                ws.cell(row=row_idx, column=4).number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
                ws.cell(row=row_idx, column=5, value=row['Rentabilidad %'])
                ws.cell(row=row_idx, column=5).number_format = '0.00"%"'
            
            # GRÁFICO - TOP PRODUCTOS
            chart_start_col = 10  # Columna J
            chart1 = BarChart()
            chart1.title = "Top 10 Productos por Utilidad"
            chart1.y_axis.title = "Utilidad ($)"
            chart1.x_axis.title = "Producto"
            
            data_ref = Reference(ws, min_col=7, min_row=start_row+1, max_row=start_row+1+len(df_productos))
            cats_ref = Reference(ws, min_col=2, min_row=start_row+2, max_row=start_row+1+len(df_productos))
            chart1.add_data(data_ref, titles_from_data=True)
            chart1.set_categories(cats_ref)
            chart1.height = 10
            chart1.width = 20
            
            ws.add_chart(chart1, f'J{start_row}')
            
            # GRÁFICO - PROVEEDORES
            chart2 = PieChart()
            chart2.title = "Utilidad por Proveedor"
            
            data_ref2 = Reference(ws, min_col=4, min_row=start_row_prov+1, max_row=start_row_prov+1+len(df_proveedores))
            cats_ref2 = Reference(ws, min_col=1, min_row=start_row_prov+2, max_row=start_row_prov+1+len(df_proveedores))
            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref2)
            chart2.height = 10
            chart2.width = 12
            
            ws.add_chart(chart2, f'J{start_row_prov}')
            
            # Autoajustar columnas
            for col_idx, column in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        # Ignorar celdas fusionadas
                        if hasattr(cell, 'value') and cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"\n✅ Reporte de utilidades mensuales guardado: {filename}")
        print(f"   Hojas creadas: {len([m for m in months_info if monthly_data.get(m['name'])])}")


# ==================== FUNCIÓN PRINCIPAL ====================

APP_ID = "7879366340316093"
CLIENT_SECRET = "TiCHBJv4QfA4Kk67xqvso3KFVlUDbsmg"
EXCEL_FILE = "CONSOLIDADO PUBLICACIONES.xlsx"

# ==================== CONFIGURACIÓN DE MODO TEST ====================
# Cambiar a False para analizar TODAS las publicaciones
TEST_MODE = False
TEST_LIMIT = 50  # Número de publicaciones a analizar en modo test

# ==================== CONFIGURACIÓN DE OPTIMIZACIONES ====================
USE_PARALLEL = True      # Paralelización (10 productos simultáneos)
ONLY_ACTIVE = False       # Solo productos activos (excluye pausados/cerrados)
# =======================================================================

print("\n" + "="*80)
print("MONITOR DE STOCK - MERCADO LIBRE CHILE")
print("Analisis basado en ventas de ultimos 30 dias")
if TEST_MODE:
    print(f"⚠ MODO PRUEBA ACTIVADO - Analizando {TEST_LIMIT} publicaciones")
if USE_PARALLEL:
    print("⚡ PARALELIZACIÓN ACTIVADA")
if ONLY_ACTIVE:
    print("⚡ FILTRO ACTIVO - Solo productos activos")
print("="*80)

monitor = MercadoLibreStockMonitor(APP_ID, CLIENT_SECRET, EXCEL_FILE)

# Autenticación
if not monitor.load_tokens():
    print("\nNecesitas autorizar la aplicacion (solo una vez)")
    input("\nPresiona ENTER para comenzar...")
    
    monitor.get_authorization_url()
    
    input("\nCuando tengas el codigo, presiona ENTER...")
    code = input("\nPega el codigo aqui: ").strip()
    
    if not code:
        print("\nNo ingresaste codigo")
    
    if not monitor.get_access_token(code):
        print("\nError en autenticacion")
else:
    print("\nTokens cargados")

# Obtener user info
print("\nObteniendo informacion de usuario...")
monitor.get_user_info()

# Análisis
print("\nIniciando analisis...")
if TEST_MODE:
    print(f"Esto puede tomar unos minutos ({TEST_LIMIT} publicaciones)...\n")
else:
    print("Esto puede tomar varios minutos (todas las publicaciones)...\n")

df = monitor.analyze_stock_needs(
    test_mode=TEST_MODE, 
    test_limit=TEST_LIMIT,
    use_parallel=USE_PARALLEL,
    only_active=ONLY_ACTIVE
)

only_report = True
# Reportes
if df is not None and not df.empty:
    # Reporte de Reposición
    monitor.generate_report(df)

    if only_report:
        pass
    else:
        # Reporte de Valorización de Inventario
        monitor.generate_inventory_report(df)
        
        # Reporte de Análisis Flex
        print("\n" + "="*80)
        print("GENERANDO REPORTE DE ANÁLISIS FLEX")
        print("="*80)
        df_flex = monitor.analyze_flex_viability(
            test_mode=TEST_MODE, 
            test_limit=TEST_LIMIT,
            only_active=ONLY_ACTIVE
        )
        if df_flex is not None and not df_flex.empty:
            monitor.generate_flex_report(df_flex)
        else:
            print("\nNo hay datos suficientes para análisis Flex")
        
        # Reporte de Utilidades Mensuales
        print("\n" + "="*80)
        print("GENERANDO REPORTE DE UTILIDADES MENSUALES")
        print("="*80)
        monthly_data, months_info = monitor.analyze_monthly_profits(
            test_mode=TEST_MODE, 
            test_limit=TEST_LIMIT, 
            months_back=6,
            only_active=ONLY_ACTIVE
        )
        if monthly_data:
            monitor.generate_monthly_profits_report(monthly_data, months_info)
        else:
            print("\nNo hay datos suficientes para análisis mensual")
elif df is not None:
    print("\nNo hay productos que necesiten reposicion")

print("\n" + "="*80)
print("Proceso completado")
print("="*80 + "\n")

# ==================== NOTIFICACIÓN SONORA ====================
def reproducir_sonido_finalizacion():
    """Reproduce un sonido para notificar que el proceso ha terminado"""
    try:
        # Intenta usar winsound (Windows)
        import winsound
        # Reproduce un sonido del sistema
        winsound.MessageBeep(winsound.MB_ICONASTERISK)
        print("🔔 Notificación sonora reproducida (Windows)")
    except ImportError:
        try:
            # Para Mac/Linux, intenta usar el comando del sistema
            import os
            import platform
            
            sistema = platform.system()
            if sistema == "Darwin":  # macOS
                os.system('afplay /System/Library/Sounds/Glass.aiff')
                print("🔔 Notificación sonora reproducida (macOS)")
            elif sistema == "Linux":
                os.system('paplay /usr/share/sounds/freedesktop/stereo/complete.oga 2>/dev/null || beep 2>/dev/null || echo -e "\a"')
                print("🔔 Notificación sonora reproducida (Linux)")
        except Exception as e:
            # Si todo falla, solo imprime un mensaje
            print("🔔 ¡Proceso completado! (No se pudo reproducir sonido)")
    except Exception as e:
        print("🔔 ¡Proceso completado! (No se pudo reproducir sonido)")

reproducir_sonido_finalizacion()